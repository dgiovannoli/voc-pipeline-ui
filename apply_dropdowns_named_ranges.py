import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation

IN_FILE = "Demo Analyst Workbook_THEMES_READY_FIXED2.xlsx"
OUT_FILE = "Demo Analyst Workbook_THEMES_READY_NAMED.xlsx"
SHEET_NAME = "Themes"
LISTS_SHEET = "Lists"

VALIDATION_OPTIONS = [
    "Validated",
    "Needs Revision",
    "Rejected",
    "Pending Review",
]

SECTION_OPTIONS = [
    "Win Drivers",
    "Loss Factors",
    "Competitive Intelligence",
    "Implementation Insights",
    "Buyer Profile",
]


def ensure_lists_sheet(wb: Workbook):
    if LISTS_SHEET in wb.sheetnames:
        ws = wb[LISTS_SHEET]
        # Clear existing
        for row in ws["A1:B100"]:
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(LISTS_SHEET)
    # Write lists
    for i, v in enumerate(VALIDATION_OPTIONS, start=1):
        ws.cell(row=i, column=1, value=v)
    for i, v in enumerate(SECTION_OPTIONS, start=1):
        ws.cell(row=i, column=2, value=v)
    # Hide sheet
    ws.sheet_state = "hidden"
    # Define named ranges
    wb.create_named_range("ValidationOptions", ws, f"$A$1:$A${len(VALIDATION_OPTIONS)}")
    wb.create_named_range("SectionOptions", ws, f"$B$1:$B${len(SECTION_OPTIONS)}")


def apply_dropdowns(wb: Workbook):
    ws = wb[SHEET_NAME]
    # Remove existing validations to avoid conflicts
    ws.data_validations.dataValidation = []
    # Create validations referencing named ranges
    dv_val = DataValidation(type="list", formula1="=ValidationOptions", allow_blank=True, showDropDown=True)
    dv_sec = DataValidation(type="list", formula1="=SectionOptions", allow_blank=True, showDropDown=True)
    ws.add_data_validation(dv_val)
    ws.add_data_validation(dv_sec)

    applied = 0
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().upper().startswith("THEME DECISION:"):
            # N11 (decision row): column N is 14
            dv_val.add(ws.cell(row=r, column=14))
            # N12 (row below): column N is 14
            sec_row = r + 1
            dv_sec.add(ws.cell(row=sec_row, column=14))
            # Optional label in B
            if not ws.cell(row=sec_row, column=2).value:
                ws.cell(row=sec_row, column=2, value="Report Section:")
            applied += 1
    return applied


def main():
    if not os.path.exists(IN_FILE):
        raise FileNotFoundError(IN_FILE)
    wb = load_workbook(IN_FILE)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found")

    ensure_lists_sheet(wb)
    count = apply_dropdowns(wb)
    wb.save(OUT_FILE)
    print(f"✅ Applied named-range dropdowns to {count} theme blocks → {OUT_FILE}")


if __name__ == "__main__":
    main() 