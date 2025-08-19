import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font

IN_FILE = "Demo Analyst Workbook_THEMES_TWO_DECISIONS.xlsx"
OUT_FILE = "Demo Analyst Workbook_THEMES_READY.xlsx"
SHEET_NAME = "Themes"

VALIDATION_OPTIONS = [
    "Validated",
    "Needs Revision",
    "Rejected",
    "Pending Review",
]

SECTION_OPTIONS = [
    "Win Drivers",
    "Loss Factors",
    "Competitive Intelligence",
    "Implementation Insights",
    "Buyer Profile",
]


def main():
    if not os.path.exists(IN_FILE):
        raise FileNotFoundError(IN_FILE)
    wb = load_workbook(IN_FILE)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found")
    ws = wb[SHEET_NAME]

    # Prepare data validations
    val_csv = ",".join(VALIDATION_OPTIONS)
    sec_csv = ",".join(SECTION_OPTIONS)
    dv_val = DataValidation(type="list", formula1=f'"{val_csv}"', allow_blank=True, showErrorMessage=True)
    dv_sec = DataValidation(type="list", formula1=f'"{sec_csv}"', allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv_val)
    ws.add_data_validation(dv_sec)

    decision_rows = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().upper().startswith("THEME DECISION:"):
            decision_rows.append(r)

    labeled = 0
    for idx, r in enumerate(decision_rows):
        # Ensure dropdown in N (14) on decision row
        dv_val.add(ws.cell(row=r, column=14).coordinate)
        # The row below (N+1 row) should already have section dropdown from previous step, ensure label in B and dropdown exists
        section_row = r + 1 + idx  # account for previously inserted extra rows shifting after prior insertions
        # Protect against overflow
        if section_row <= ws.max_row:
            # Label in column B
            ws.cell(row=section_row, column=2, value="Report Section:").font = Font(bold=True)
            # Ensure dropdown in N on section row
            dv_sec.add(ws.cell(row=section_row, column=14).coordinate)
            labeled += 1

    wb.save(OUT_FILE)
    print(f"✅ Finalized: added Validation dropdowns to N on decision rows and labeled 'Report Section' with dropdown on the row below. → {OUT_FILE}")


if __name__ == "__main__":
    main() 