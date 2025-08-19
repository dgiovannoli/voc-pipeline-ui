import os
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC_WORKBOOK = "Demo Analyst Workbook.xlsx"
OUT_WORKBOOK = "Demo Analyst Workbook_THEMES_FROM_TEMPLATE.xlsx"
THEMES_SHEET_NAME = "Themes"
WIN_SHEET_KEY = "win drivers"

SECTION_SHEET_MATCHERS = [
    ("win drivers", "Win Drivers Section"),
    ("loss factors", "Loss Factors Section"),
    ("competitive intelligence", "Competitive Intelligence"),
    ("implementation", "Implementation Insights"),
]


def find_sheet_by_key(wb, key: str) -> Worksheet:
    key = key.lower()
    for name in wb.sheetnames:
        if key in name.lower():
            return wb[name]
    raise RuntimeError(f"Sheet containing '{key}' not found")


def find_theme_blocks(ws: Worksheet) -> List[Tuple[int, int]]:
    """Return list of (start_row, end_row) for each theme block, inclusive of decision row."""
    blocks: List[Tuple[int, int]] = []
    start = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.startswith("theme_"):
            start = r
        if isinstance(a, str) and a.strip().upper().startswith("THEME DECISION:") and start is not None:
            end = r
            blocks.append((start, end))
            start = None
    return blocks


def get_template_rows(ws_win: Worksheet) -> Tuple[int, int, int, int]:
    """Find header rows count, first theme header row, first quote row, decision row in Win Drivers sheet."""
    header_rows_end = 1
    header_row = None
    decision_row = None
    for r in range(1, ws_win.max_row + 1):
        a = ws_win.cell(row=r, column=1).value
        if isinstance(a, str) and a.startswith("theme_") and header_row is None:
            header_row = r
            break
        header_rows_end = r
    # Find decision row for the first block
    for r in range(header_row, ws_win.max_row + 1):
        a = ws_win.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().upper().startswith("THEME DECISION:"):
            decision_row = r
            break
    # First quote row is the next non-empty row after header_row if column F has text, else header_row+1
    quote_row = header_row + 1
    return header_rows_end, header_row, quote_row, decision_row


def clone_style_from(src_cell, dst_cell):
    # Create new style objects to avoid StyleProxy issues
    sfont = src_cell.font
    dst_cell.font = Font(name=sfont.name, size=sfont.size, bold=sfont.bold, italic=sfont.italic,
                         vertAlign=sfont.vertAlign, underline=sfont.underline, strike=sfont.strike,
                         color=sfont.color)
    sfill = src_cell.fill
    dst_cell.fill = PatternFill(fill_type=sfill.fill_type, fgColor=sfill.fgColor, bgColor=sfill.bgColor)
    salign = src_cell.alignment
    dst_cell.alignment = Alignment(horizontal=salign.horizontal, vertical=salign.vertical,
                                   wrap_text=salign.wrap_text, shrink_to_fit=salign.shrink_to_fit,
                                   indent=salign.indent, text_rotation=salign.text_rotation)
    sb = src_cell.border
    def copy_side(side: Side) -> Side:
        return Side(style=side.style, color=side.color)
    dst_cell.border = Border(left=copy_side(sb.left), right=copy_side(sb.right),
                             top=copy_side(sb.top), bottom=copy_side(sb.bottom))
    dst_cell.number_format = src_cell.number_format


def copy_row_style(src_ws: Worksheet, src_row: int, dst_ws: Worksheet, dst_row: int, max_col: int):
    for c in range(1, max_col + 1):
        clone_style_from(src_ws.cell(row=src_row, column=c), dst_ws.cell(row=dst_row, column=c))


def clear_data_after(ws: Worksheet, start_row: int):
    if ws.max_row > start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def main():
    if not os.path.exists(SRC_WORKBOOK):
        raise FileNotFoundError(SRC_WORKBOOK)
    wb = load_workbook(SRC_WORKBOOK)

    ws_win = find_sheet_by_key(wb, WIN_SHEET_KEY)
    header_end, tmpl_header_row, tmpl_quote_row, tmpl_decision_row = get_template_rows(ws_win)
    max_col = ws_win.max_column

    # Create Themes by duplicating Win Drivers section to preserve sheet-level formatting
    if THEMES_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[THEMES_SHEET_NAME])
    ws_out = wb.copy_worksheet(ws_win)
    ws_out.title = THEMES_SHEET_NAME

    # Clear all data rows after the header rows
    clear_data_after(ws_out, header_end + 1)

    # Prepare a list of blocks from all section sheets
    all_blocks = []
    for key, _ in SECTION_SHEET_MATCHERS:
        try:
            ws = find_sheet_by_key(wb, key)
        except RuntimeError:
            continue
        blocks = find_theme_blocks(ws)
        if not blocks:
            continue
        all_blocks.append((ws, blocks))

    # Append blocks using Win Drivers styles
    current = header_end + 1
    for src_ws, blocks in all_blocks:
        for (start, end) in blocks:
            # 1) Header row from src start -> apply header style
            ws_out.insert_rows(current)
            copy_row_style(ws_win, tmpl_header_row, ws_out, current, max_col)
            for c in range(1, max_col + 1):
                ws_out.cell(row=current, column=c).value = src_ws.cell(row=start, column=c).value
            current += 1
            # 2) Intermediate quote rows from start+1 .. end-1 -> apply quote style
            for r in range(start + 1, end):
                ws_out.insert_rows(current)
                copy_row_style(ws_win, tmpl_quote_row, ws_out, current, max_col)
                for c in range(1, max_col + 1):
                    ws_out.cell(row=current, column=c).value = src_ws.cell(row=r, column=c).value
                current += 1
            # 3) Decision row (end) -> apply decision style
            ws_out.insert_rows(current)
            copy_row_style(ws_win, tmpl_decision_row, ws_out, current, max_col)
            for c in range(1, max_col + 1):
                ws_out.cell(row=current, column=c).value = src_ws.cell(row=end, column=c).value
            current += 1
            # 4) Add one spacer row (copy style from quote row but empty)
            ws_out.insert_rows(current)
            copy_row_style(ws_win, tmpl_quote_row, ws_out, current, max_col)
            for c in range(1, max_col + 1):
                ws_out.cell(row=current, column=c).value = None
            current += 1

    wb.save(OUT_WORKBOOK)
    print(f"✅ Built Themes from Win Drivers template: {OUT_WORKBOOK}")


if __name__ == "__main__":
    main() 