import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

IN_FILE = "Demo Analyst Workbook_THEMES_FROM_TEMPLATE.xlsx"
OUT_FILE = "Demo Analyst Workbook_THEMES_TWO_DECISIONS.xlsx"
SHEET_NAME = "Themes"

SECTION_OPTIONS = [
    "Win Drivers",
    "Loss Factors",
    "Competitive Intelligence",
    "Implementation Insights",
    "Buyer Profile",
]


def clone_style(src_cell, dst_cell):
    sfont = src_cell.font
    dst_cell.font = Font(name=sfont.name, size=sfont.size, bold=sfont.bold, italic=sfont.italic,
                         vertAlign=sfont.vertAlign, underline=sfont.underline, strike=sfont.strike,
                         color=sfont.color)
    sfill = src_cell.fill
    dst_cell.fill = PatternFill(fill_type=sfill.fill_type, fgColor=sfill.fgColor, bgColor=sfill.bgColor)
    salign = src_cell.alignment
    dst_cell.alignment = Alignment(horizontal=salign.horizontal, vertical=salign.vertical,
                                   wrap_text=salign.wrap_text, shrink_to_fit=salign.shrink_to_fit,
                                   indent=salign.indent, text_rotation=salign.text_rotation)
    sb = src_cell.border
    def copy_side(side: Side) -> Side:
        return Side(style=side.style, color=side.color)
    dst_cell.border = Border(left=copy_side(sb.left), right=copy_side(sb.right),
                             top=copy_side(sb.top), bottom=copy_side(sb.bottom))
    dst_cell.number_format = src_cell.number_format


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int):
    for c in range(1, max_col + 1):
        clone_style(ws.cell(row=src_row, column=c), ws.cell(row=dst_row, column=c))


def main():
    if not os.path.exists(IN_FILE):
        raise FileNotFoundError(IN_FILE)
    wb = load_workbook(IN_FILE)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found")
    ws = wb[SHEET_NAME]

    max_col = ws.max_column

    # Collect decision row indices first
    decision_rows = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().upper().startswith("THEME DECISION:"):
            decision_rows.append(r)

    # Add validation list
    options_csv = ",".join(SECTION_OPTIONS)
    dv = DataValidation(type="list", formula1=f'"{options_csv}"', allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)

    # Insert a new row after each decision row; adjust offset as rows shift
    offset = 0
    for r in decision_rows:
        insert_at = r + 1 + offset
        ws.insert_rows(insert_at)
        # Copy style from decision row into inserted row
        copy_row_style(ws, r + offset, insert_at, max_col)
        # Add dropdown at column N (14)
        dv.add(ws.cell(row=insert_at, column=14).coordinate)
        offset += 1

    wb.save(OUT_FILE)
    print(f"✅ Added second decision row with dropdown in column N for {len(decision_rows)} themes → {OUT_FILE}")


if __name__ == "__main__":
    main() 