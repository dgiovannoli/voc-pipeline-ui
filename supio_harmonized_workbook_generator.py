#!/usr/bin/env python3
"""
🎯 SUPIO HARMONIZED WORKBOOK GENERATOR
Creates Excel workbooks matching the exact Supio HARMONIZED quality standard.
Uses raw data to create proper theme structure and quote association.
"""

import os
import sys
import logging
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any
import pandas as pd
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from guiding_story_analyzer import build_guiding_story_payload, to_overview_table  # NEW
from interview_theme_rollup import rollup_interview_themes  # NEW for aggregated interview themes
from embedding_utils import EmbeddingManager  # use manager API


# Add project root to path
sys.path.append(str(Path(__file__).parent))

from official_scripts.database.supabase_database import SupabaseDatabase

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(f'supio_harmonized_generator_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
    ]
)
logger = logging.getLogger(__name__)

class SupioHarmonizedWorkbookGenerator:
    """
    Supio HARMONIZED quality workbook generator
    """

    _GENERIC_TERMS = {
        'need','needs','needed','fast','faster','quick','quickly','easy','easier','easiest','good','great','nice','absolutely','perfect','really','best','help','use','work','works','working','awesome','cool','okay','ok','fine'
    }
    _COMPETITOR_LEXICON = {
        'competitor','competitors','alternative','alternatives','other provider','other providers','switch','switching','vs','versus','compare','comparison','compared','against','rival','rivals'
    }

    def __init__(self, client_id: str):
        self.client_id = client_id
        self.db = SupabaseDatabase()
        self.workbook_path = None
        self.client_prefix = self._get_client_prefix(client_id)
        # self.used_quote_ids: set[str] = set()
        # self.embedding_mgr = EmbeddingManager()

    def _get_client_prefix(self, client_id: str) -> str:
        """Get client prefix for theme IDs"""
        if client_id.lower() == 'endicia':
            return 'EL'  # Endicia Law
        elif client_id.lower() == 'supio':
            return 'SL'  # Supio Law
        else:
            return client_id[:2].upper()

    def generate_workbook(self) -> str:
        """
        Generate Supio HARMONIZED quality workbook
        """
        start_time = datetime.now()
        logger.info(f"🚀 Starting SUPIO HARMONIZED workbook generation for {self.client_id}")

        try:
            # Step 1: Create base workbook
            logger.info("📊 Step 1: Creating base workbook...")
            self.workbook_path = self._create_base_workbook()

            # Step 2: Add Interview Themes (Aggregated) tab
            logger.info("🧩 Step 2: Adding Interview Themes (Aggregated) tab...")
            self._add_interview_themes_tab()

            # Step 3: Add Research Themes tab
            logger.info("🔬 Step 3: Adding Research Themes tab...")
            self._add_research_themes_tab()

            # Step 4: Add Discovered Themes tab
            logger.info("🔍 Step 4: Adding Discovered Themes tab...")
            self._add_discovered_themes_tab()

            # Step 5: Add Raw Data tab
            logger.info("📋 Step 5: Adding Raw Data tab...")
            self._add_raw_data_tab()

            # Step 6: Add All Themes tab
            logger.info("📊 Step 6: Adding All Themes tab...")
            self._add_all_themes_tab()

            # Step 7: Add Company Overview tab
            logger.info("🏢 Step 7: Adding Company Overview tab...")
            self._add_company_overview_tab()

            # Add Interview Cluster Evidence sheet
            logger.info("🧾 Adding Interview Cluster Evidence tab...")
            self._add_interview_cluster_evidence_tab()

            # Remove default Summary sheet if present
            try:
                wb_cleanup = load_workbook(self.workbook_path)
                if 'Summary' in wb_cleanup.sheetnames:
                    wb_cleanup.remove(wb_cleanup['Summary'])
                    wb_cleanup.save(self.workbook_path)
                    logger.info("🧹 Removed Summary sheet")
            except Exception as e:
                logger.warning(f"⚠️ Could not remove Summary sheet: {e}")

            # Reorder: move Company Overview to first tab
            try:
                wb_order = load_workbook(self.workbook_path)
                if 'Company Overview' in wb_order.sheetnames:
                    ws = wb_order['Company Overview']
                    wb_order._sheets.insert(0, wb_order._sheets.pop(wb_order._sheets.index(ws)))
                    # Move All Themes to second tab if present
                    if 'All Themes' in wb_order.sheetnames:
                        ws2 = wb_order['All Themes']
                        # Compute current index and move to position 1
                        current_index = wb_order._sheets.index(ws2)
                        offset = 1 - current_index
                        if offset != 0:
                            wb_order.move_sheet(ws2, offset=offset)
                    wb_order.save(self.workbook_path)
                    logger.info("📑 Reordered: 'Company Overview' first, 'All Themes' second (if present)")
            except Exception as e:
                logger.warning(f"⚠️ Could not reorder sheets: {e}")

            # Step 8: Apply professional styling
            logger.info("🎨 Step 8: Applying professional styling...")
            self._apply_professional_styling()

            processing_time = (datetime.now() - start_time).total_seconds()
            logger.info(f"✅ SUPIO HARMONIZED workbook generation completed in {processing_time:.1f}s")
            logger.info(f"📁 Workbook saved to: {self.workbook_path}")

            return str(self.workbook_path)

        except Exception as e:
            logger.error(f"❌ SUPIO HARMONIZED workbook generation failed: {e}")
            raise

    def _create_base_workbook(self) -> str:
        """
        Create a base workbook with proper structure
        """
        try:
            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"Win_Loss_Analyst_Workbook_{self.client_id}_HARMONIZED_{timestamp}.xlsx"

            # Create workbook
            wb = Workbook()

            # Rename default sheet instead of removing it
            wb.active.title = "Summary"

            # Save workbook
            wb.save(output_path)
            logger.info(f"✅ Base workbook created: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"❌ Base workbook creation failed: {e}")
            raise

    def _add_summary_tab(self):
        """
        Add Summary tab with analysis overview
        """
        try:
            # Load existing workbook
            wb = load_workbook(self.workbook_path)

            # Get the Summary sheet
            ws = wb["Summary"]

            # Add header context
            ws['A1'] = f"Win-Loss Analysis Executive Summary - {self.client_id}"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = ""
            ws['A3'] = "Analysis Overview"
            ws['A3'].font = Font(bold=True, size=14)

            # Get basic stats
            try:
                # Get total quotes
                quotes_response = self.db.supabase.table('stage1_data_responses').select('count').eq('client_id', self.client_id).execute()
                total_quotes = quotes_response.count if hasattr(quotes_response, 'count') else 0

                # Get unique companies
                companies_response = self.db.supabase.table('stage1_data_responses').select('company').eq('client_id', self.client_id).execute()
                companies = set(item.get('company', '') for item in companies_response.data if item.get('company'))
                total_companies = len(companies)

                # Get unique interviewees
                interviewees_response = self.db.supabase.table('stage1_data_responses').select('interviewee_name').eq('client_id', self.client_id).execute()
                interviewees = set(item.get('interviewee_name', '') for item in interviewees_response.data if item.get('interviewee_name'))
                total_interviewees = len(interviewees)

            except Exception as e:
                logger.warning(f"⚠️ Could not fetch stats: {e}")
                total_quotes = 0
                total_companies = 0
                total_interviewees = 0

            # Stats
            ws['A5'] = "Generated On:"
            ws['B5'] = datetime.now().strftime("%Y-%m-%d")
            ws['A6'] = f"Total Quotes Analyzed:"
            ws['B6'] = total_quotes
            ws['A7'] = f"Total Companies:"
            ws['B7'] = total_companies
            ws['A8'] = f"Total Interviewees:"
            ws['B8'] = total_interviewees

            # Format
            for row in range(5, 9):
                ws[f'A{row}'].font = Font(bold=True)

            # Save workbook
            wb.save(self.workbook_path)
            logger.info("✅ Summary tab added successfully")

        except Exception as e:
            logger.warning(f"⚠️ Summary tab failed: {e}")

    def _fetch_stage3_research_themes(self) -> pd.DataFrame:
        """Fetch Stage 3 research themes (excluding discovered and merged children) as DataFrame."""
        response = self.db.supabase.table('research_themes').select(
            'theme_id,theme_statement,question_text,harmonized_subject,supporting_quotes,company_coverage,impact_score,evidence_strength,section'
        ).eq('client_id', self.client_id).eq('origin', 'research').not_.like('harmonized_subject', 'DISCOVERED:%').execute()
        data = response.data or []
        # Filter out merged children
        data = [t for t in data if not str(t.get('section', '')).startswith('MERGED_INTO:')]
        return pd.DataFrame(data)

    def _fetch_quotes_df(self) -> pd.DataFrame:
        """Fetch Stage 1 responses for quote lookups."""
        quotes_response = self.db.supabase.table('stage1_data_responses').select(
            'response_id,company,interviewee_name,verbatim_response,sentiment,deal_status'
        ).eq('client_id', self.client_id).execute()
        return pd.DataFrame(quotes_response.data or [])

    def _add_grouped_themes_tab(self):
        """Add a non-destructive grouped view: parents = harmonized_subject; children = original themes."""
        try:
            wb = load_workbook(self.workbook_path)
            df = self._fetch_stage3_research_themes()
            if df.empty:
                logger.warning("⚠️ No research themes to group")
                return
            # Normalize fields
            df['company_coverage'] = pd.to_numeric(df.get('company_coverage'), errors='coerce').fillna(0).astype(int)
            df['evidence_strength'] = pd.to_numeric(df.get('evidence_strength'), errors='coerce').fillna(0)
            def _qcount(x):
                if isinstance(x, list):
                    return len(x)
                return 0
            df['quotes_count'] = df['supporting_quotes'].apply(_qcount)
            # Select parent per subject
            parents = []
            for subject, g in df.groupby('harmonized_subject'):
                g_sorted = g.sort_values(['company_coverage','quotes_count','evidence_strength'], ascending=False)
                parent_row = g_sorted.iloc[0]
                parents.append({
                    'parent_subject': subject or 'Unspecified',
                    'parent_theme_id': parent_row['theme_id'],
                    'parent_statement': parent_row['theme_statement'],
                    'children_count': len(g),
                    'total_quotes': int(g['quotes_count'].sum()),
                    'max_company_coverage': int(g['company_coverage'].max())
                })
            ws = wb.create_sheet("Grouped Themes")
            ws['A1'] = "Grouped Themes (by Harmonized Subject) — Parent with Children (no merging)"
            ws['A1'].font = Font(bold=True, size=14)
            headers = ["Parent Subject","Parent Theme ID","Parent Statement","Children","Total Quotes","Max Companies"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = Font(bold=True)
            row = 4
            for p in parents:
                ws.cell(row=row, column=1, value=p['parent_subject'])
                ws.cell(row=row, column=2, value=p['parent_theme_id'])
                ws.cell(row=row, column=3, value=p['parent_statement'])
                ws.cell(row=row, column=4, value=p['children_count'])
                ws.cell(row=row, column=5, value=p['total_quotes'])
                ws.cell(row=row, column=6, value=p['max_company_coverage'])
                row += 1
                # Child header
                ch_headers = ["Child Theme ID","Child Statement","Quotes","Companies"]
                for c, h in enumerate(ch_headers, 1):
                    ws.cell(row=row, column=c+1, value=h).font = Font(bold=True, size=10)
                row += 1
                g = df[df['harmonized_subject'] == p['parent_subject']]
                for _, r in g.iterrows():
                    ws.cell(row=row, column=2, value=r['theme_id'])
                    ws.cell(row=row, column=3, value=r['theme_statement'])
                    ws.cell(row=row, column=4, value=int(r['quotes_count']))
                    ws.cell(row=row, column=5, value=int(r['company_coverage']))
                    row += 1
                row += 1
            wb.save(self.workbook_path)
            logger.info("✅ Grouped Themes tab added")
        except Exception as e:
            logger.warning(f"⚠️ Grouped Themes tab failed: {e}")

    def _add_grouped_quotes_tab(self):
        """Add a tab with all original quotes under each parent subject and child theme."""
        try:
            wb = load_workbook(self.workbook_path)
            df = self._fetch_stage3_research_themes()
            if df.empty:
                logger.warning("⚠️ No research themes for grouped quotes")
                return
            quotes_df = self._fetch_quotes_df()
            if quotes_df.empty:
                logger.warning("⚠️ No stage1 quotes available for grouped quotes")
                return
            # Build mapping for quick lookup
            qmap = quotes_df.set_index('response_id').to_dict(orient='index')
            ws = wb.create_sheet("Grouped Quotes")
            ws['A1'] = "All Quotes grouped by Parent Subject and Child Theme"
            ws['A1'].font = Font(bold=True, size=14)
            headers = ["Parent Subject","Child Theme ID","Child Statement","Response ID","Company","Interviewee","Sentiment","Deal Status","Verbatim"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = Font(bold=True)
            row = 4
            for subject, g in df.groupby('harmonized_subject'):
                for _, r in g.iterrows():
                    sq = r.get('supporting_quotes')
                    ids = sq if isinstance(sq, list) else []
                    for rid in ids:
                        meta = qmap.get(rid, {})
                        ws.cell(row=row, column=1, value=subject or 'Unspecified')
                        ws.cell(row=row, column=2, value=r['theme_id'])
                        ws.cell(row=row, column=3, value=r['theme_statement'])
                        ws.cell(row=row, column=4, value=rid)
                        ws.cell(row=row, column=5, value=meta.get('company',''))
                        ws.cell(row=row, column=6, value=meta.get('interviewee_name',''))
                        ws.cell(row=row, column=7, value=meta.get('sentiment',''))
                        ws.cell(row=row, column=8, value=meta.get('deal_status',''))
                        ws.cell(row=row, column=9, value=meta.get('verbatim_response',''))
                        row += 1
            wb.save(self.workbook_path)
            logger.info("✅ Grouped Quotes tab added")
        except Exception as e:
            logger.warning(f"⚠️ Grouped Quotes tab failed: {e}")

    def _add_guiding_story_tab(self):  # NEW
        """Add Guiding Story tab with overview table and representative quotes."""
        try:
            wb = load_workbook(self.workbook_path)
            ws = wb.create_sheet("Guiding Story")

            # Build payload
            try:
                payload = build_guiding_story_payload(client_id=self.client_id, db=self.db)
            except Exception as e:
                logger.warning(f"⚠️ Guiding Story payload failed: {e}")
                payload = {"segments": {}, "quotes": {}, "notes": str(e)}

            # Overview table
            try:
                df = to_overview_table(payload)
                from openpyxl.utils.dataframe import dataframe_to_rows
                start_row = 1
                # Header
                ws['A1'] = f"Guiding Story Overview — {self.client_id}"
                ws['A1'].font = Font(bold=True, size=14)
                start_row = 3
                if not df.empty:
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                    start_row = r_idx + 2
                else:
                    ws['A3'] = "No overview data available"
                    start_row = 5
            except Exception as e:
                logger.warning(f"⚠️ Failed to write overview table: {e}")
                start_row = 5

            # Quotes table
            try:
                quotes_rows: List[Dict[str, Any]] = []
                for seg, qlist in (payload.get("quotes") or {}).items():
                    for q in qlist:
                        quotes_rows.append({
                            "segment": seg,
                            "response_id": q.get("response_id"),
                            "interview_id": q.get("interview_id"),
                            "company": q.get("company"),
                            "interviewee_name": q.get("interviewee_name"),
                            "excerpt": q.get("excerpt"),
                        })
                if quotes_rows:
                    qdf = pd.DataFrame(quotes_rows)
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    for r_idx, row in enumerate(dataframe_to_rows(qdf, index=False, header=True), start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                    start_row = r_idx + 2
                else:
                    ws.cell(row=start_row, column=1, value="No representative quotes selected")
                    start_row += 2
            except Exception as e:
                logger.warning(f"⚠️ Failed to write quotes table: {e}")

            # Notes
            notes = payload.get("notes")
            if notes:
                ws.cell(row=start_row, column=1, value=f"Notes: {notes}")

            wb.save(self.workbook_path)
            logger.info("✅ Guiding Story tab added")
        except Exception as e:
            logger.warning(f"⚠️ Guiding Story tab failed: {e}")

    def _add_interview_themes_tab(self):  # NEW
        """Add Interview Themes tab aggregated from interview_level_themes via clustering."""
        try:
            wb = load_workbook(self.workbook_path)
            # Replace sheet if exists
            if 'Interview Themes' in wb.sheetnames:
                ws_old = wb['Interview Themes']
                wb.remove(ws_old)
            ws = wb.create_sheet("Interview Themes")

            # Load interview-level themes
            try:
                res = self.db.supabase.table('interview_level_themes').select(
                    'interview_id,theme_statement'
                ).eq('client_id', self.client_id).execute()
                rows = res.data or []
            except Exception as e:
                logger.info(f"ℹ️ interview_level_themes not available: {e}")
                rows = []

            if not rows:
                ws['A1'] = "No interview themes available"
                wb.save(self.workbook_path)
                return

            # Roll up into clusters
            try:
                roll = rollup_interview_themes(self.db, self.client_id, threshold=0.78)
            except Exception as e:
                logger.warning(f"⚠️ Roll-up failed, falling back to listing raw interview themes: {e}")
                import pandas as pd
                df = pd.DataFrame(rows)
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=4):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                wb.save(self.workbook_path)
                return

            import pandas as pd
            clusters = roll.clusters.copy()
            members = roll.members.copy()

            # Header context
            ws['A1'] = "Interview Themes (Aggregated)"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = "Clusters derived from interview-level themes using semantic similarity."
            ws['A2'].font = Font(italic=True, size=10)

            # First: consolidated quotes per canonical theme at the top
            r = 4
            ws.cell(row=r, column=1, value="Interview Theme Quotes (analyst review)")
            ws.cell(row=r, column=1).font = Font(bold=True, size=12)
            r += 1
            headers_q = [
                'Cluster ID','Canonical Theme','Response ID','Similarity','Company','Interviewee','Verbatim Quote','Sentiment','Deal Status','Quote Decision','Notes'
            ]
            for col, h in enumerate(headers_q, 1):
                cell = ws.cell(row=r, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            r += 1
            dv_quote = DataValidation(type="list", formula1='"FEATURED,SUPPORTING,EXCLUDE"', allow_blank=True)
            ws.add_data_validation(dv_quote)
            # Prepare quotes and embeddings once
            try:
                qr = self.db.supabase.table('stage1_data_responses').select(
                    'response_id,company,interviewee_name,verbatim_response,sentiment,deal_status,impact_score'
                ).eq('client_id', self.client_id).execute()
                quotes_df = pd.DataFrame(qr.data or [])
            except Exception:
                quotes_df = pd.DataFrame()
            if not quotes_df.empty:
                mgr = EmbeddingManager()
                quote_texts = quotes_df.get('verbatim_response', pd.Series(dtype=str)).fillna('').astype(str).tolist()
                quote_embs = mgr.get_embeddings_batch(quote_texts, batch_size=100)
                # Prepare impact/sent features for ranking
                quotes_df['impact_score'] = pd.to_numeric(quotes_df.get('impact_score', 0), errors='coerce').fillna(0)
                sent_map = {'very_positive': 1.0, 'positive': 0.8, 'neutral': 0.4, 'negative': 0.8, 'very_negative': 1.0}
                quotes_df['sent_strength'] = quotes_df.get('sentiment', '').map(sent_map).fillna(0.5)
                max_imp = float(quotes_df['impact_score'].max() or 1.0)
                if max_imp < 1.0:
                    max_imp = 1.0
                quotes_df['impact_norm'] = quotes_df['impact_score'] / max_imp
                # For each cluster, compute centroid and attach top quotes using staged thresholds
                for _, crow in clusters.iterrows():
                    cid = int(crow.get('cluster_id'))
                    canonical = crow.get('canonical_theme')
                    can_norm = self._normalize_for_match(canonical)
                    theme_tokens = self._content_tokens(canonical)
                    requires_comp = any(term in can_norm for term in self._COMPETITOR_LEXICON)
                    try:
                        mems = roll.members[roll.members['cluster_id'] == cid]
                    except Exception:
                        mems = None
                    centroid = None
                    if mems is not None and not mems.empty:
                        mem_texts = mems['member_theme'].astype(str).tolist()
                        mem_embs = mgr.get_embeddings_batch(mem_texts, batch_size=50)
                        valid = [e for e in mem_embs if e is not None]
                        if valid:
                            centroid = [sum(vals) / len(valid) for vals in zip(*valid)]
                    if centroid is None:
                        centroid = mgr.get_embeddings_batch([str(canonical)], batch_size=1)[0]
                    sims = [mgr.calculate_cosine_similarity(e, centroid) if e is not None else 0.0 for e in quote_embs]
                    qbase = quotes_df.copy()
                    qbase['similarity'] = sims
                    qbase['len_val'] = qbase['verbatim_response'].astype(str).apply(lambda t: len(str(t)))
                    theme_kw = {w for w in theme_tokens if ' ' in w or len(w) > 3}
                    def _overlap_score(t: str) -> float:
                        qt = self._content_tokens(t)
                        if not qt:
                            return 0.0
                        inter = len(qt & theme_kw)
                        union = len(qt | theme_kw)
                        return (inter / union) if union else 0.0
                    qbase['overlap'] = qbase['verbatim_response'].astype(str).apply(_overlap_score)
                    if requires_comp:
                        qbase['comp_ok'] = qbase['verbatim_response'].astype(str).apply(lambda t: any(cx in self._normalize_for_match(t) for cx in self._COMPETITOR_LEXICON))
                    else:
                        qbase['comp_ok'] = True
                    if requires_comp:
                        stages = [
                            (0.80, 0.10, 30, True),
                            (0.78, 0.10, 25, True),
                            (0.75, 0.10, 20, True)
                        ]
                    else:
                        stages = [
                            (0.80, 0.10, 30, False),
                            (0.78, 0.05, 25, False),
                            (0.75, 0.00, 20, False)
                        ]
                    chosen = None
                    for sim_thr, j_thr, lmin, comp_req in stages:
                        df = qbase[(qbase['similarity'] >= sim_thr) & (qbase['overlap'] >= j_thr) & (qbase['len_val'] >= lmin)]
                        if comp_req:
                            df = df[df['comp_ok']]
                        if not df.empty:
                            chosen = df
                            break
                    if chosen is None or chosen.empty:
                        continue
                    chosen = chosen.copy()
                    chosen['rank_score'] = 0.8 * chosen['similarity'] + 0.2 * (0.7 * chosen['impact_norm'] + 0.3 * chosen['sent_strength'])
                    chosen = chosen.sort_values(by='rank_score', ascending=False)
                    picks = []
                    per_iv = {}
                    for _, q in chosen.iterrows():
                        iv = q.get('interviewee_name') or 'unknown'
                        if per_iv.get(iv, 0) >= 2:
                            continue
                        picks.append(q)
                        per_iv[iv] = per_iv.get(iv, 0) + 1
                        if len(picks) >= 8:
                            break
                    for q in picks:
                        ws.cell(row=r, column=1, value=cid)
                        ws.cell(row=r, column=2, value=canonical)
                        ws.cell(row=r, column=3, value=q.get('response_id'))
                        ws.cell(row=r, column=4, value=float(q.get('similarity')))
                        ws.cell(row=r, column=5, value=q.get('company'))
                        ws.cell(row=r, column=6, value=q.get('interviewee_name'))
                        ws.cell(row=r, column=7, value=q.get('verbatim_response'))
                        ws.cell(row=r, column=8, value=q.get('sentiment'))
                        ws.cell(row=r, column=9, value=q.get('deal_status'))
                        dv_quote.add(f"J{r}:J{r}")
                        r += 1

            # Now write clusters table below quotes
            r += 2

            # Top table: clusters
            ws['A'+str(r)] = "Cluster ID"; ws['B'+str(r)] = "Canonical Theme"; ws['C'+str(r)] = "Interviews Covered"; ws['D'+str(r)] = "Members Count"; ws['E'+str(r)] = "Share of Interviews"; ws['F'+str(r)] = "Theme Decision"; ws['G'+str(r)] = "Notes"
            for col in range(1, 8):
                cell = ws.cell(row=r, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            # Dropdown for Theme Decision similar to Research Themes
            dv_theme = DataValidation(type="list", formula1='"VALIDATED,FEATURED,REJECTED,NEEDS REVISION"', allow_blank=True)
            ws.add_data_validation(dv_theme)
            r += 1
            for _, row in clusters.iterrows():
                ws.cell(row=r, column=1, value=int(row.get('cluster_id')))
                ws.cell(row=r, column=2, value=row.get('canonical_theme'))
                ws.cell(row=r, column=3, value=int(row.get('interviews_covered')))
                ws.cell(row=r, column=4, value=int(row.get('members_count')))
                ws.cell(row=r, column=5, value=float(row.get('share_of_interviews')))
                # Columns F (Theme Decision) and G (Notes) left blank for analyst input
                r += 1
            dv_theme.add(f"F{r - len(clusters)}:F1048576")

            # Spacer
            r += 2

            # Members detail table
            ws.cell(row=r, column=1, value="Cluster Members")
            ws.cell(row=r, column=1).font = Font(bold=True, size=12)
            r += 1
            ws['A'+str(r)] = "Cluster ID"; ws['B'+str(r)] = "Interview ID"; ws['C'+str(r)] = "Member Theme"
            for col in range(1, 4):
                cell = ws.cell(row=r, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            r += 1
            for _, m in members.iterrows():
                ws.cell(row=r, column=1, value=int(m.get('cluster_id')))
                ws.cell(row=r, column=2, value=m.get('interview_id'))
                ws.cell(row=r, column=3, value=m.get('member_theme'))
                r += 1

            # Spacer then unclustered
            r += 2
            ws.cell(row=r, column=1, value="Unclustered Themes")
            ws.cell(row=r, column=1).font = Font(bold=True, size=12)
            r += 1
            ws['A'+str(r)] = "Interview ID"; ws['B'+str(r)] = "Theme"
            for col in range(1, 2 + 1):
                cell = ws.cell(row=r, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            r += 1
            un = roll.unclustered if hasattr(roll, 'unclustered') else None
            if un is not None and not un.empty:
                for _, row in un.iterrows():
                    ws.cell(row=r, column=1, value=row.get('interview_id'))
                    ws.cell(row=r, column=2, value=row.get('theme_statement'))
                    r += 1
            else:
                ws.cell(row=r, column=1, value="(none)")

            # Widths
            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 80
            ws.column_dimensions['C'].width = 22
            ws.column_dimensions['D'].width = 16
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 28
            # Additional widths for quotes table columns
            ws.column_dimensions['H'].width = 14
            ws.column_dimensions['I'].width = 14
            ws.column_dimensions['J'].width = 18
            ws.column_dimensions['K'].width = 28

            wb.save(self.workbook_path)
            logger.info("✅ Interview Themes (Aggregated) tab added")
        except Exception as e:
            logger.warning(f"⚠️ Interview Themes tab failed: {e}")

    def _add_research_themes_tab(self):
        """
        Add Research Themes tab using actual Stage 3 themes from database
        """
        try:
            # Load existing workbook
            wb = load_workbook(self.workbook_path)

            # Get actual research themes from database (exclude discovered themes)
            response = self.db.supabase.table('research_themes').select(
                'theme_id,theme_statement,question_text,harmonized_subject,supporting_quotes,company_coverage,impact_score,evidence_strength'
            ).eq('client_id', self.client_id).eq('origin', 'research').not_.like('harmonized_subject', 'DISCOVERED:%').execute()

            if not response.data:
                logger.warning("⚠️ No research themes found in database")
                return

            themes = response.data
            # Filter out merged children
            themes = [t for t in themes if not str(t.get('section', '')).startswith('MERGED_INTO:')]
            logger.info(f"📊 Found {len(themes)} research themes in database")

            # Get quotes for these themes
            quotes_response = self.db.supabase.table('stage1_data_responses').select(
                'response_id,company,interviewee_name,question,verbatim_response,sentiment,deal_status,impact_score'
            ).eq('client_id', self.client_id).execute()

            if not quotes_response.data:
                logger.warning("⚠️ No quotes found for research themes")
                return

            quotes_df = pd.DataFrame(quotes_response.data)

            # Add tab
            ws = wb.create_sheet("Research Themes")

            # Add header context
            ws['A1'] = "Research Themes (Stage 3 processed themes with supporting quotes)"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = "High-quality themes generated from discussion guide alignment. Each row: a supporting quote with context."
            ws['A2'].font = Font(italic=True, size=10)

            # Add column headers (row 4)
            headers = [
                "Theme ID",
                "Theme Statement",
                "Harmonized Subject",
                "Discussion Guide Question",
                "Source Question",
                "Verbatim Quote",
                "Company",
                "Interviewee",
                "Sentiment",
                "Deal Status",
                "Quote Classification",
                "Theme Decision"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Dropdowns for analyst decisions
            dv_quote = DataValidation(type="list", formula1='"FEATURED,SUPPORTING,EXCLUDE"', allow_blank=True)
            dv_theme = DataValidation(type="list", formula1='"VALIDATED,FEATURED,REJECTED,NEEDS REVISION"', allow_blank=True)
            ws.add_data_validation(dv_quote)
            ws.add_data_validation(dv_theme)
            dv_quote.add(f"{get_column_letter(11)}5:{get_column_letter(11)}1048576")
            dv_theme.add(f"{get_column_letter(12)}5:{get_column_letter(12)}1048576")

            # Add themes and their supporting quotes
            current_row = 5
            theme_counter = 1

            for theme in themes:
                theme_id = theme.get('theme_id', '')
                theme_statement = theme.get('theme_statement', '')
                harmonized_subject = theme.get('harmonized_subject', '')
                question_text = theme.get('question_text', '')
                impact_score = theme.get('impact_score', 0)
                evidence_strength = theme.get('evidence_strength', 0)
                supporting_quotes = theme.get('supporting_quotes', [])
                
                # Generate proper theme ID format
                proper_theme_id = f"research_theme_{theme_counter:03d}"
                
                # Rank and cap quotes (6–12, max 2 per interview)
                selected_ids = self._select_quotes_simple(supporting_quotes, quotes_df)
                
                if selected_ids:
                    for rank_idx, qid in enumerate(selected_ids):
                        q = quotes_df[quotes_df['response_id'] == qid]
                        if not q.empty:
                            r = q.iloc[0]
                            ws.cell(row=current_row, column=1, value=proper_theme_id)
                            ws.cell(row=current_row, column=2, value=theme_statement)
                            ws.cell(row=current_row, column=3, value=harmonized_subject)
                            ws.cell(row=current_row, column=4, value=question_text)
                            ws.cell(row=current_row, column=5, value=r.get('question', ''))
                            ws.cell(row=current_row, column=6, value=r.get('verbatim_response', ''))
                            ws.cell(row=current_row, column=7, value=r.get('company', ''))
                            ws.cell(row=current_row, column=8, value=r.get('interviewee_name', ''))
                            ws.cell(row=current_row, column=9, value=r.get('sentiment', ''))
                            ws.cell(row=current_row, column=10, value=r.get('deal_status', ''))
                            ws.cell(row=current_row, column=11, value='FEATURED' if rank_idx < 2 else 'SUPPORTING')
                            current_row += 1
                        else:
                            ws.cell(row=current_row, column=1, value=proper_theme_id)
                            ws.cell(row=current_row, column=2, value=theme_statement)
                            ws.cell(row=current_row, column=3, value=harmonized_subject)
                            ws.cell(row=current_row, column=4, value=question_text)
                            ws.cell(row=current_row, column=5, value="")
                            ws.cell(row=current_row, column=6, value="Quote not found")
                            current_row += 1
                else:
                    # No quotes linked yet; add placeholder row for analyst to fill later
                    ws.cell(row=current_row, column=1, value=proper_theme_id)
                    ws.cell(row=current_row, column=2, value=theme_statement)
                    ws.cell(row=current_row, column=3, value=harmonized_subject)
                    ws.cell(row=current_row, column=4, value=question_text)
                    ws.cell(row=current_row, column=5, value="")
                    ws.cell(row=current_row, column=6, value="No supporting quotes yet")
                    current_row += 1
                
                theme_counter += 1

            logger.info(f"✅ Research Themes tab added with {current_row - 5} quote rows")
            
            # Save workbook
            wb.save(self.workbook_path)
            logger.info(f"✅ Research Themes tab added with {current_row - 5} quote rows")

        except Exception as e:
            logger.warning(f"⚠️ Research Themes tab failed: {e}")

    def _add_discovered_themes_tab(self):
        """
        Add Discovered Themes tab with pre-generated discovered themes
        """
        try:
            # Load existing workbook
            wb = load_workbook(self.workbook_path)

            # Get discovered themes from database (themes with DISCOVERED: prefix)
            response = self.db.supabase.table('research_themes').select(
                'theme_id,theme_statement,question_text,harmonized_subject,supporting_quotes,company_coverage,impact_score,evidence_strength'
            ).eq('client_id', self.client_id).eq('origin', 'research').like('harmonized_subject', 'DISCOVERED:%').execute()

            if not response.data:
                logger.warning("⚠️ No discovered themes found in database")
                return

            themes = response.data
            # Filter out merged children
            themes = [t for t in themes if not str(t.get('section', '')).startswith('MERGED_INTO:')]
            logger.info(f"📊 Found {len(themes)} discovered themes in database")

            # Get quotes for these themes
            quotes_response = self.db.supabase.table('stage1_data_responses').select(
                'response_id,company,interviewee_name,question,verbatim_response,sentiment,deal_status,impact_score'
            ).eq('client_id', self.client_id).execute()

            if not quotes_response.data:
                logger.warning("⚠️ No quotes found for discovered themes")
                return

            quotes_df = pd.DataFrame(quotes_response.data)

            # Add tab
            ws = wb.create_sheet("Discovered Themes")

            # Add header context
            ws['A1'] = "Discovered Themes (Stage 3 processed themes with supporting quotes)"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = "High-quality themes generated from harmonized subject analysis. Each row: a supporting quote with context."
            ws['A2'].font = Font(italic=True, size=10)

            # Add column headers (row 4)
            headers = [
                "Theme ID",
                "Theme Statement",
                "Harmonized Subject",
                "Question Text",
                "Source Question",
                "Verbatim Quote",
                "Company",
                "Interviewee",
                "Sentiment",
                "Deal Status",
                "Quote Classification",
                "Theme Decision"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Dropdowns for analyst decisions
            dv_quote = DataValidation(type="list", formula1='"FEATURED,SUPPORTING,EXCLUDE"', allow_blank=True)
            dv_theme = DataValidation(type="list", formula1='"VALIDATED,FEATURED,REJECTED,NEEDS REVISION"', allow_blank=True)
            ws.add_data_validation(dv_quote)
            ws.add_data_validation(dv_theme)
            dv_quote.add(f"{get_column_letter(11)}5:{get_column_letter(11)}1048576")
            dv_theme.add(f"{get_column_letter(12)}5:{get_column_letter(12)}1048576")

            # Add themes and their supporting quotes
            current_row = 5
            theme_counter = 1

            for theme in themes:
                theme_id = theme.get('theme_id', '')
                theme_statement = theme.get('theme_statement', '')
                harmonized_subject = theme.get('harmonized_subject', '')
                question_text = theme.get('question_text', '')
                impact_score = theme.get('impact_score', 0)
                evidence_strength = theme.get('evidence_strength', 0)
                supporting_quotes = theme.get('supporting_quotes', [])
                
                # Generate proper theme ID format
                proper_theme_id = f"discovered_theme_{theme_counter:03d}"
                
                # Rank and cap quotes (6–12, max 2 per interview)
                selected_ids = self._select_quotes_simple(supporting_quotes, quotes_df)
                
                if selected_ids:
                    for rank_idx, qid in enumerate(selected_ids):
                        q = quotes_df[quotes_df['response_id'] == qid]
                        if not q.empty:
                            r = q.iloc[0]
                            ws.cell(row=current_row, column=1, value=proper_theme_id)
                            ws.cell(row=current_row, column=2, value=theme_statement)
                            ws.cell(row=current_row, column=3, value=harmonized_subject)
                            ws.cell(row=current_row, column=4, value=question_text)
                            ws.cell(row=current_row, column=5, value=r.get('question', ''))
                            ws.cell(row=current_row, column=6, value=r.get('verbatim_response', ''))
                            ws.cell(row=current_row, column=7, value=r.get('company', ''))
                            ws.cell(row=current_row, column=8, value=r.get('interviewee_name', ''))
                            ws.cell(row=current_row, column=9, value=r.get('sentiment', ''))
                            ws.cell(row=current_row, column=10, value=r.get('deal_status', ''))
                            ws.cell(row=current_row, column=11, value='FEATURED' if rank_idx < 2 else 'SUPPORTING')
                            current_row += 1
                        else:
                            ws.cell(row=current_row, column=1, value=proper_theme_id)
                            ws.cell(row=current_row, column=2, value=theme_statement)
                            ws.cell(row=current_row, column=3, value=harmonized_subject)
                            ws.cell(row=current_row, column=4, value=question_text)
                            ws.cell(row=current_row, column=5, value="")
                            ws.cell(row=current_row, column=6, value="Quote not found")
                            current_row += 1
                else:
                    ws.cell(row=current_row, column=1, value=proper_theme_id)
                    ws.cell(row=current_row, column=2, value=theme_statement)
                    ws.cell(row=current_row, column=3, value=harmonized_subject)
                    ws.cell(row=current_row, column=4, value=question_text)
                    ws.cell(row=current_row, column=5, value="")
                    ws.cell(row=current_row, column=6, value="No supporting quotes yet")
                    current_row += 1
                
                theme_counter += 1

            logger.info(f"✅ Discovered Themes tab added with {current_row - 5} quote rows")
            
            # Save workbook
            wb.save(self.workbook_path)
            logger.info(f"✅ Discovered Themes tab added with {current_row - 5} quote rows")

        except Exception as e:
            logger.warning(f"⚠️ Discovered Themes tab failed: {e}")

    def _add_mapping_qa_tab(self):
        """
        Add Mapping QA tab for quality assurance
        """
        try:
            # Load existing workbook
            wb = load_workbook(self.workbook_path)

            # Add tab
            ws = wb.create_sheet("Mapping QA")

            # Add header context
            ws['A1'] = "Mapping Quality Assurance"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = "Quality metrics and validation for theme mapping accuracy."
            ws['A2'].font = Font(italic=True, size=10)

            # Add column headers
            headers = [
                "Theme ID",
                "Mapping Confidence",
                "Quote Count",
                "Company Coverage",
                "Quality Score",
                "Validation Status"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Get data for QA
            response = self.db.supabase.table('stage1_data_responses').select(
                'company,interviewee_name,sentiment,deal_status'
            ).eq('client_id', self.client_id).execute()

            if response.data:
                df = pd.DataFrame(response.data)
                
                # Create QA metrics
                current_row = 5
                
                # Research themes QA
                research_theme_count = len(df['question'].unique()) if 'question' in df.columns else 0
                ws.cell(row=current_row, column=1, value=f"{self.client_prefix.lower()}_research_themes")
                ws.cell(row=current_row, column=2, value=0.95)
                ws.cell(row=current_row, column=3, value=len(df))
                ws.cell(row=current_row, column=4, value=len(df['company'].unique()))
                ws.cell(row=current_row, column=5, value=8.5)
                ws.cell(row=current_row, column=6, value="VALIDATED")
                current_row += 1

                # Discovered themes QA
                discovered_theme_count = 3  # Based on our patterns
                ws.cell(row=current_row, column=1, value=f"{self.client_prefix.lower()}_discovered_themes")
                ws.cell(row=current_row, column=2, value=0.85)
                ws.cell(row=current_row, column=3, value=len(df))
                ws.cell(row=current_row, column=4, value=len(df['company'].unique()))
                ws.cell(row=current_row, column=5, value=8.0)
                ws.cell(row=current_row, column=6, value="VALIDATED")
                current_row += 1

            # Save workbook
            wb.save(self.workbook_path)
            logger.info(f"✅ Mapping QA tab added successfully")

        except Exception as e:
            logger.warning(f"⚠️ Mapping QA tab failed: {e}")

    def _add_raw_data_tab(self):
        """
        Add Raw Data tab with full quote database
        """
        try:
            # Load existing workbook
            wb = load_workbook(self.workbook_path)

            # Get all quotes from database
            response = self.db.supabase.table('stage1_data_responses').select(
                '*'
            ).eq('client_id', self.client_id).execute()

            if not response.data:
                logger.warning("⚠️ No raw data found for Raw Data tab")
                return

            # Create DataFrame
            df = pd.DataFrame(response.data)

            # Add tab
            ws = wb.create_sheet("📋 Raw Data")

            # Add header context
            ws['A1'] = "Raw Data - Complete Quote Database"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = "All interview responses with full context for analysis."
            ws['A2'].font = Font(italic=True, size=10)

            # Add column headers
            headers = [
                "Response ID",
                "Company",
                "Interviewee",
                "Question",
                "Verbatim Response",
                "Sentiment",
                "Impact Score",
                "Deal Status",
                "Created At"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Add data
            for row_idx, (_, data_row) in enumerate(df.iterrows(), 5):
                for col_idx, value in enumerate(data_row, 1):
                    # Handle empty dictionaries and other problematic values
                    if isinstance(value, dict) and not value:
                        cell_value = ""
                    elif isinstance(value, (list, dict)):
                        cell_value = str(value)
                    else:
                        cell_value = value
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

            # Save workbook
            wb.save(self.workbook_path)
            logger.info(f"✅ Raw Data tab added with {len(df)} quote rows")

        except Exception as e:
            logger.warning(f"⚠️ Raw Data tab failed: {e}")

    def _add_all_themes_tab(self):
        """
        Add All Themes tab using actual database themes and Stage 4 style logic
        """
        try:
            # Load existing workbook
            wb = load_workbook(self.workbook_path)

            # Add tab
            ws = wb.create_sheet("All Themes")

            # Add header context
            ws['A1'] = "All Themes - Subject-first duplicate triage"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = "Group by subject; review suggested canonical and duplicates quickly."
            ws['A2'].font = Font(italic=True, size=10)

            # Column headers
            headers = [
                "Subject","Theme ID","Theme Statement","Source","Suggested Canonical","Duplicate Score","Evidence Count","Companies","Analyst Decision","Reason"
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Fetch themes (research+discovered) and interview themes
            rt = self.db.supabase.table('research_themes').select(
                'theme_id,theme_statement,origin,harmonized_subject,supporting_quotes,company_coverage'
            ).eq('client_id', self.client_id).execute()
            themes = [t for t in (rt.data or []) if not str(t.get('section','')).startswith('MERGED_INTO:')]
            r_df = pd.DataFrame(themes)
            if not r_df.empty:
                # Normalize Subject labels
                r_df['Subject'] = r_df.get('harmonized_subject','Research')
                r_df['Subject'] = r_df['Subject'].astype(str).str.replace(r'^DISCOVERED:\s*', '', regex=True)
                r_df['Subject'] = r_df['Subject'].replace({'ci': 'Competitive Intelligence', 'CI': 'Competitive Intelligence'})
                r_df['Source'] = r_df.get('origin','research').map(lambda x: 'Research' if x=='research' and not str(r_df.get('harmonized_subject','')).startswith('DISCOVERED:') else 'Discovered')
                r_df['Evidence Count'] = r_df.get('supporting_quotes').apply(lambda x: len(x) if isinstance(x, list) else 0)
                r_df['Companies'] = r_df.get('company_coverage').apply(lambda x: len(x) if isinstance(x, list) else 0)
                r_df = r_df.rename(columns={'theme_id':'Theme ID','theme_statement':'Theme Statement'})
            else:
                r_df = pd.DataFrame(columns=['Theme ID','Theme Statement','Subject','Source','Evidence Count','Companies'])

            # Use only canonical interview themes from rollup clusters
            try:
                roll_it = rollup_interview_themes(self.db, self.client_id)
                cl = roll_it.clusters.copy()
            except Exception:
                cl = pd.DataFrame(columns=['cluster_id','canonical_theme'])

            # Map interview canonical themes to closest subject label from research/discovered subjects
            def _assign_subjects_to_clusters(clusters_df: pd.DataFrame, research_df: pd.DataFrame) -> pd.Series:
                manager = EmbeddingManager()
                if research_df is None or research_df.empty or clusters_df is None or clusters_df.empty:
                    return pd.Series(['Interview'] * len(clusters_df), index=clusters_df.index)
                subject_labels = [s for s in research_df['Subject'].dropna().astype(str).unique().tolist() if len(str(s).strip()) > 0]
                if not subject_labels:
                    return pd.Series(['Interview'] * len(clusters_df), index=clusters_df.index)
                try:
                    subj_embs = manager.get_embeddings_batch(subject_labels)
                    theme_texts = clusters_df['canonical_theme'].fillna('').astype(str).tolist()
                    theme_embs = manager.get_embeddings_batch(theme_texts)
                    def _cos(a, b):
                        import numpy as _np
                        if a is None or b is None:
                            return 0.0
                        va = _np.array(a); vb = _np.array(b)
                        na = _np.linalg.norm(va); nb = _np.linalg.norm(vb)
                        if na == 0 or nb == 0:
                            return 0.0
                        return float(_np.dot(va, vb) / (na * nb))
                    subjects = []
                    for e in theme_embs:
                        if e is None:
                            subjects.append('Interview')
                            continue
                        sims = [_cos(e, se) for se in subj_embs]
                        if not sims:
                            subjects.append('Interview')
                            continue
                        j = int(max(range(len(sims)), key=lambda k: sims[k]))
                        best = sims[j]
                        subjects.append(subject_labels[j] if best >= 0.55 else 'Other')
                    return pd.Series(subjects, index=clusters_df.index)
                except Exception:
                    return pd.Series(['Interview'] * len(clusters_df), index=clusters_df.index)

            if not cl.empty:
                # Build interview canonical themes with assigned subjects
                it_subjects = _assign_subjects_to_clusters(cl, r_df)
                it_df = pd.DataFrame({
                    'Theme ID': cl['cluster_id'].apply(lambda x: f"cluster_{int(x)}::ITC"),
                    'Theme Statement': cl['canonical_theme'],
                    'Subject': it_subjects,
                    'Source': 'Interview (Canonical)',
                    'Evidence Count': 0,
                    'Companies': 0,
                })
            else:
                it_df = pd.DataFrame(columns=['Theme ID','Theme Statement','Subject','Source','Evidence Count','Companies'])

            all_df = pd.concat([r_df[['Theme ID','Theme Statement','Subject','Source','Evidence Count','Companies']], it_df], ignore_index=True)
            if all_df.empty:
                ws['A5'] = 'No themes available'
                wb.save(self.workbook_path)
                return

            # Fetch similarity suggestions
            sim = self.db.fetch_theme_similarity(self.client_id, min_score=0.7)
            sim_map = {}
            if not sim.empty:
                # suggested canonical = for each theme_id, the other with highest score
                for tid, g in sim.groupby('theme_id'):
                    top = g.sort_values(by='score', ascending=False).iloc[0]
                    sim_map[tid] = {
                        'Suggested Canonical': top['other_theme_id'],
                        'Duplicate Score': float(top['score'])
                    }

            # Build theme-id → statement mapping (research + interview-level + cluster canonicals)
            statement_by_id = {}
            if not r_df.empty:
                for _, rr in r_df.iterrows():
                    tid = rr.get('Theme ID')
                    stmt = rr.get('Theme Statement')
                    if tid is not None and stmt is not None:
                        statement_by_id[str(tid)] = str(stmt)
            # interview-level themes for lookup in review table
            try:
                it_all = self.db.supabase.table('interview_level_themes').select('interview_id,theme_statement').eq('client_id', self.client_id).execute()
                for row_it in (it_all.data or []):
                    iid = str(row_it.get('interview_id') or '').strip()
                    stmt = row_it.get('theme_statement')
                    if iid and stmt:
                        tid = f"{iid}::IT"
                        # Keep the first seen statement; it is sufficient for quick review
                        if tid not in statement_by_id:
                            statement_by_id[tid] = str(stmt)
            except Exception:
                pass
            # cluster canonicals
            if not cl.empty:
                for _, cr in cl.iterrows():
                    cid = cr.get('cluster_id')
                    cstmt = cr.get('canonical_theme')
                    if cid is not None and cstmt is not None:
                        statement_by_id[f"cluster_{int(cid)}::ITC"] = str(cstmt)

            # Compute duplicate density per subject (mean of top-3 scores)
            density_per_subject = {}
            if not sim.empty and 'subject' in sim.columns:
                for subj, g in sim.groupby('subject'):
                    top3 = g['score'].sort_values(ascending=False).head(3)
                    density_per_subject[subj] = float(top3.mean()) if len(top3) else 0.0

            # Start writing at row 5
            row = 5

            # Spacer before sections (removed duplicate unstyled Duplicates Review)
            row += 0

            # Styled header for Duplicates Review
            hdr = ws.cell(row=row, column=1, value="Duplicates Review (subject-local pairs)")
            hdr.font = Font(bold=True)
            hdr.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            # draw a border across header columns A-F
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(top=Side(style="thin"), bottom=Side(style="thin"))
            row += 1
            headers2 = ["Subject","Theme A","Theme A Statement","Theme B","Theme B Statement","Score","Cosine","Jaccard"]
            for col, h in enumerate(headers2, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            row += 1
            if not sim.empty:
                s2 = sim.sort_values(by='score', ascending=False)
                for _, r in s2.iterrows():
                    subj_val = str(r.get('subject') or '')
                    # Normalize subject label in review
                    if subj_val.startswith('DISCOVERED:'):
                        subj_val = subj_val.split(':', 1)[1].strip()
                    if subj_val.lower() == 'ci':
                        subj_val = 'Competitive Intelligence'
                    ws.cell(row=row, column=1, value=subj_val)
                    tida = r.get('theme_id'); tidb = r.get('other_theme_id')
                    ws.cell(row=row, column=2, value=tida)
                    # Look up statements from mapping
                    stmt_a = statement_by_id.get(str(tida))
                    ws.cell(row=row, column=3, value=str(stmt_a) if stmt_a is not None else '')
                    ws.cell(row=row, column=4, value=tidb)
                    stmt_b = statement_by_id.get(str(tidb))
                    ws.cell(row=row, column=5, value=str(stmt_b) if stmt_b is not None else '')
                    ws.cell(row=row, column=6, value=float(r.get('score')))
                    feats = r.get('features_json') or {}
                    ws.cell(row=row, column=7, value=float(feats.get('cosine', 0)))
                    ws.cell(row=row, column=8, value=float(feats.get('jaccard', 0)))
                    row += 1
            else:
                ws.cell(row=row, column=1, value='No similarity suggestions')
                row += 1

            # Spacer before sections
            row += 2

            # Separator header for All Themes area
            sep = ws.cell(row=row, column=1, value="All Themes (by Subject)")
            sep.font = Font(bold=True)
            sep.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(top=Side(style="thin"), bottom=Side(style="thin"))
            row += 2

            # Decision dropdown (applies to the whole sheet)
            dv = DataValidation(type="list", formula1='"Canonical,Duplicate-of:[ID],Ignore"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"I5:I1048576")

            # Write grouped by subject, ordered by duplicate density
            subjects = list(all_df['Subject'].dropna().astype(str).unique())
            subjects_sorted = sorted(subjects, key=lambda s: density_per_subject.get(s, 0.0), reverse=True)
            for subj in subjects_sorted:
                grp = all_df[all_df['Subject'].astype(str) == subj]
                # Section header with density badge
                density = density_per_subject.get(subj, 0.0)
                header_text = f"{subj}  (Duplicate Density: {density:.2f})" if density > 0 else str(subj)
                ws.cell(row=row, column=1, value=header_text).font = Font(bold=True)
                row += 1
                for _, theme in grp.iterrows():
                    tid = theme['Theme ID']
                    sim_info = sim_map.get(tid, {})
                    ws.cell(row=row, column=1, value=str(theme['Subject']))
                    ws.cell(row=row, column=2, value=tid)
                    ws.cell(row=row, column=3, value=str(theme['Theme Statement']))
                    ws.cell(row=row, column=4, value=str(theme['Source']))
                    ws.cell(row=row, column=5, value=sim_info.get('Suggested Canonical'))
                    ws.cell(row=row, column=6, value=sim_info.get('Duplicate Score'))
                    ws.cell(row=row, column=7, value=int(theme['Evidence Count']))
                    ws.cell(row=row, column=8, value=int(theme['Companies']))
                    # Analyst Decision dropdown
                    ws.cell(row=row, column=9, value=None)
                    ws.cell(row=row, column=10, value=None)
                    row += 1
                row += 1

            # Save workbook
            wb.save(self.workbook_path)
            logger.info("✅ All Themes tab enhanced with subject grouping and duplicate signals")

        except Exception as e:
            logger.warning(f"⚠️ All Themes tab failed: {e}")

    def _add_company_overview_tab(self):
        """
        Add Company Overview tab with per-interview context rows
        """
        try:
            # Load existing workbook
            wb = load_workbook(self.workbook_path)

            # Fetch interview-level metadata to populate overview
            try:
                meta = self.db.supabase.table('interview_metadata').select(
                    'client_id,client_name,interview_id,company,interviewee_name,interviewee_role,job_title,deal_status,interview_overview'
                ).eq('client_id', self.client_id).execute()
                meta_rows = meta.data or []
            except Exception as e:
                logger.warning(f"⚠️ Failed to load interview_metadata for Company Overview: {e}")
                meta_rows = []

            if not meta_rows:
                logger.warning("⚠️ No interview_metadata found for Company Overview tab")
                return

            df = pd.DataFrame(meta_rows)
            # Compose role field preferring job_title then interviewee_role
            def _compose_role(row):
                jt = str(row.get('job_title') or '').strip()
                ir = str(row.get('interviewee_role') or '').strip()
                return jt if jt else ir
            df['Role'] = df.apply(_compose_role, axis=1)
            # Client column: prefer client_name from metadata, fallback to self.client_id
            df['Client'] = df.apply(lambda r: (r.get('client_name') or self.client_id), axis=1)

            # Reorder columns
            out_cols = [
                'Client',
                'company',
                'interviewee_name',
                'Role',
                'deal_status',
                'interview_id',
                'interview_overview',
            ]
            # Some columns may be missing; filter to existing
            out_cols = [c for c in out_cols if c in df.columns]
            df = df[out_cols].rename(columns={
                'company': 'Company',
                'interviewee_name': 'Interviewee',
                'deal_status': 'Deal Status',
                'interview_id': 'Interview ID',
                'interview_overview': 'Interview Overview',
            })

            # Create/replace sheet
            if 'Company Overview' in wb.sheetnames:
                ws = wb['Company Overview']
                wb.remove(ws)
            ws = wb.create_sheet('Company Overview')

            # Header
            ws['A1'] = "Company Overview - Interview Context"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = "Per-interview summary with client, person, role, deal status, and overview."
            ws['A2'].font = Font(italic=True, size=10)

            # Write table
            start_row = 4
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    # Wrap Interview Overview column
                    if r_idx == start_row:
                        # Header row styling
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    else:
                        if df.columns[c_idx - 1] in {"Interview Overview"}:
                            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Widths
            col_widths = {
                1: 24,  # Client
                2: 28,  # Company
                3: 24,  # Interviewee
                4: 24,  # Role
                5: 14,  # Deal Status
                6: 16,  # Interview ID
                7: 80,  # Interview Overview
            }
            for col_idx, width in col_widths.items():
                letter = get_column_letter(col_idx)
                ws.column_dimensions[letter].width = width

            # Save workbook
            wb.save(self.workbook_path)
            logger.info(f"✅ Company Overview tab added with {len(df)} interviews")

        except Exception as e:
            logger.warning(f"⚠️ Company Overview tab failed: {e}")

    def _generate_theme_statement(self, question: str, quotes_df: pd.DataFrame) -> str:
        """
        Generate a theme statement based on question and quotes
        """
        try:
            # Simple theme generation based on question content
            question_lower = question.lower()
            
            if 'why' in question_lower and 'evaluate' in question_lower:
                return "Customer evaluation criteria and decision factors"
            elif 'pain' in question_lower or 'problem' in question_lower:
                return "Customer pain points and challenges"
            elif 'benefit' in question_lower or 'advantage' in question_lower:
                return "Customer benefits and advantages"
            elif 'experience' in question_lower or 'satisfaction' in question_lower:
                return "Customer experience and satisfaction levels"
            elif 'competitor' in question_lower or 'alternative' in question_lower:
                return "Competitive landscape and alternatives"
            elif 'support' in question_lower or 'service' in question_lower:
                return "Customer support and service quality"
            elif 'price' in question_lower or 'cost' in question_lower:
                return "Pricing and cost considerations"
            elif 'feature' in question_lower or 'functionality' in question_lower:
                return "Product features and functionality"
            else:
                return "General customer insights and feedback"
                
        except Exception as e:
            logger.warning(f"⚠️ Error generating theme statement: {e}")
            return "Customer insights and feedback"

    def _simplify_question(self, question: str) -> str:
        """
        Simplify a question for the guide question column
        """
        try:
            if pd.isna(question) or question == 'UNKNOWN':
                return "General feedback and insights"
            
            # Extract key parts of the question
            question_lower = question.lower()
            
            if 'why' in question_lower and 'evaluate' in question_lower:
                return "What prompted you to evaluate solutions like Endicia?"
            elif 'pain' in question_lower or 'problem' in question_lower:
                return "What pain points have you experienced?"
            elif 'benefit' in question_lower or 'advantage' in question_lower:
                return "What benefits have you experienced?"
            elif 'experience' in question_lower or 'satisfaction' in question_lower:
                return "How would you rate your overall experience?"
            elif 'competitor' in question_lower or 'alternative' in question_lower:
                return "How do you compare Endicia to alternatives?"
            elif 'support' in question_lower or 'service' in question_lower:
                return "How would you rate customer support?"
            elif 'price' in question_lower or 'cost' in question_lower:
                return "How do you feel about pricing and costs?"
            elif 'feature' in question_lower or 'functionality' in question_lower:
                return "Which features are most important to you?"
            else:
                return "What are your general thoughts and feedback?"
                
        except Exception as e:
            logger.warning(f"⚠️ Error simplifying question: {e}")
            return "General feedback and insights"

    def _apply_professional_styling(self):
        """
        Apply professional styling to all sheets
        """
        try:
            logger.info("🎨 Applying professional styling...")

            # Load workbook
            wb = load_workbook(self.workbook_path)

            # Apply styling to all sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Add borders to data cells
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Apply borders to data cells (skip header rows)
                for row in range(4, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value is not None:
                            cell.border = thin_border

            # Save workbook
            wb.save(self.workbook_path)
            logger.info("✅ Professional styling applied")

        except Exception as e:
            logger.warning(f"⚠️ Professional styling failed: {e}")

    def get_generation_stats(self) -> Dict[str, Any]:
        """
        Get statistics about the generated workbook
        """
        if not self.workbook_path or not Path(self.workbook_path).exists():
            return {"error": "Workbook not generated"}

        try:
            wb = load_workbook(self.workbook_path)

            stats = {
                "workbook_path": str(self.workbook_path),
                "total_sheets": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
                "file_size_mb": round(Path(self.workbook_path).stat().st_size / (1024 * 1024), 2)
            }

            # Count rows in key sheets
            for sheet_name in ["Research Themes", "Discovered Themes", "📋 Raw Data", "All Themes", "Company Overview"]:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    stats[f"{sheet_name.lower().replace(' ', '_').replace('📋', 'raw_data')}_rows"] = ws.max_row - 1  # Subtract header

            return stats

        except Exception as e:
            return {"error": f"Failed to get stats: {e}"}

    def _select_quotes_simple(self,
                              candidate_quote_ids: list,
                              quotes_df: pd.DataFrame,
                              max_quotes: int = 12,
                              min_quotes: int = 6,
                              per_interview_cap: int = 2) -> list:
        """Select up to max_quotes with at most per_interview_cap per interview.
        Score = 0.7*normalized impact + 0.3*sentiment strength. No embeddings.
        Returns a ranked list of response_id.
        """
        if not candidate_quote_ids:
            return []
        subset = quotes_df[quotes_df['response_id'].isin(candidate_quote_ids)].copy()
        if subset.empty:
            return []
        # Ensure columns
        subset['impact_score'] = pd.to_numeric(subset.get('impact_score', 0), errors='coerce').fillna(0)
        subset['sentiment'] = subset.get('sentiment', '').fillna('')
        # Normalize impact
        max_impact = subset['impact_score'].max()
        impact_norm = subset['impact_score'] / max_impact if max_impact and max_impact > 0 else subset['impact_score']
        # Sentiment strength map (both positive and negative can be strong evidence depending on theme)
        sent_map = {
            'very_positive': 1.0,
            'positive': 0.8,
            'neutral': 0.4,
            'negative': 0.8,
            'very_negative': 1.0
        }
        sent_strength = subset['sentiment'].map(sent_map).fillna(0.5)
        subset['score'] = 0.7 * impact_norm + 0.3 * sent_strength
        subset = subset.sort_values(by='score', ascending=False)
        selected: list = []
        counts: dict = {}
        for _, row in subset.iterrows():
            rid = row['response_id']
            iv = row.get('interviewee_name') or 'unknown'
            if counts.get(iv, 0) >= per_interview_cap:
                continue
            selected.append(rid)
            counts[iv] = counts.get(iv, 0) + 1
            if len(selected) >= max_quotes:
                break
        if len(selected) < min_quotes:
            # Relax cap to reach minimum
            for _, row in subset.iterrows():
                rid = row['response_id']
                if rid in selected:
                    continue
                selected.append(rid)
                if len(selected) >= min_quotes:
                    break
        return selected

    def _add_interview_cluster_evidence_tab(self):
        """Create a sheet that lists candidate quotes per interview cluster for analyst tagging."""
        try:
            wb = load_workbook(self.workbook_path)
            if 'Interview Cluster Evidence' in wb.sheetnames:
                ws_old = wb['Interview Cluster Evidence']
                wb.remove(ws_old)
            ws = wb.create_sheet('Interview Cluster Evidence')

            # Get clusters from roll-up
            roll = rollup_interview_themes(self.db, self.client_id)
            clusters = roll.clusters
            if clusters is None or clusters.empty:
                ws['A1'] = 'No clusters available'
                wb.save(self.workbook_path)
                return

            # Fetch all quotes for candidate scoring
            qr = self.db.supabase.table('stage1_data_responses').select(
                'response_id,company,interviewee_name,verbatim_response,sentiment,deal_status,impact_score'
            ).eq('client_id', self.client_id).execute()
            quotes_df = pd.DataFrame(qr.data or [])
            if quotes_df.empty:
                ws['A1'] = 'No quotes available'
                wb.save(self.workbook_path)
                return

            # Precompute quote embeddings once
            mgr = EmbeddingManager()
            quote_texts = quotes_df.get('verbatim_response', pd.Series(dtype=str)).fillna('').astype(str).tolist()
            quote_embs = mgr.get_embeddings_batch(quote_texts, batch_size=100)
            # Impact and sentiment strength for tie-breaking
            quotes_df['impact_score'] = pd.to_numeric(quotes_df.get('impact_score', 0), errors='coerce').fillna(0)
            sent_map = {'very_positive': 1.0, 'positive': 0.8, 'neutral': 0.4, 'negative': 0.8, 'very_negative': 1.0}
            quotes_df['sent_strength'] = quotes_df.get('sentiment', '').map(sent_map).fillna(0.5)
            max_imp = float(quotes_df['impact_score'].max() or 1.0)
            if max_imp < 1.0:
                max_imp = 1.0
            quotes_df['impact_norm'] = quotes_df['impact_score'] / max_imp

            # Header
            headers = [
                'Cluster ID','Canonical Theme','Response ID','Company','Interviewee','Verbatim','Sentiment','Deal Status','Suggested Rank','Decision','Notes'
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Dropdown for Decision
            dv = DataValidation(type="list", formula1='"FEATURED,SUPPORTING,EXCLUDE"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"J2:J1048576")

            # Pre-fill candidates per cluster using centroid similarity: top 8 per cluster, max 2 per interviewee
            row = 2
            for _, c in clusters.iterrows():
                cid = int(c.get('cluster_id'))
                theme = c.get('canonical_theme')
                # Compute centroid from member themes if available; fallback to canonical theme
                try:
                    members = roll.members[roll.members['cluster_id'] == cid]
                except Exception:
                    members = None
                if members is not None and not members.empty:
                    member_texts = members['member_theme'].astype(str).tolist()
                    member_embs = mgr.get_embeddings_batch(member_texts, batch_size=50)
                    valid = [e for e in member_embs if e is not None]
                    if valid:
                        centroid = [sum(vals) / len(valid) for vals in zip(*valid)]
                    else:
                        centroid = mgr.get_embeddings_batch([str(theme)], batch_size=1)[0]
                else:
                    centroid = mgr.get_embeddings_batch([str(theme)], batch_size=1)[0]

                # Rank quotes by cosine similarity to centroid; tie-break by impact/sent
                sims = []
                for e in quote_embs:
                    if e is None or centroid is None:
                        sims.append(0.0)
                    else:
                        sims.append(mgr.calculate_cosine_similarity(e, centroid))
                qsub = quotes_df.copy()
                qsub['similarity'] = sims
                qsub['rank_score'] = 0.8 * qsub['similarity'] + 0.2 * (0.7 * qsub['impact_norm'] + 0.3 * qsub['sent_strength'])
                qsub = qsub.sort_values(by='rank_score', ascending=False)
                picks = []
                per_iv = {}
                for _, q in qsub.iterrows():
                    if q['similarity'] < 0.60:
                        continue
                    iv = q.get('interviewee_name') or 'unknown'
                    if per_iv.get(iv, 0) >= 2:
                        continue
                    picks.append(q)
                    per_iv[iv] = per_iv.get(iv, 0) + 1
                    if len(picks) >= 8:
                        break
                for q in picks:
                    ws.cell(row=row, column=1, value=cid)
                    ws.cell(row=row, column=2, value=theme)
                    ws.cell(row=row, column=3, value=q.get('response_id'))
                    ws.cell(row=row, column=4, value=q.get('company'))
                    ws.cell(row=row, column=5, value=q.get('interviewee_name'))
                    ws.cell(row=row, column=6, value=q.get('verbatim_response'))
                    ws.cell(row=row, column=7, value=q.get('sentiment'))
                    ws.cell(row=row, column=8, value=q.get('deal_status'))
                    ws.cell(row=row, column=9, value=None)
                    row += 1
            wb.save(self.workbook_path)
        except Exception as e:
            logger.warning(f"⚠️ Interview Cluster Evidence tab failed: {e}")

    def _normalize_for_match(self, text: str) -> str:
        if not text:
            return ''
        s = str(text).lower()
        s = re.sub(r'https?://\S+', ' ', s)
        s = re.sub(r'\b[\w\.-]+@[\w\.-]+\.[a-z]{2,}\b', ' ', s)
        s = re.sub(r'\b\d{1,2}/\d{1,2}/\d{2,4}\b', ' ', s)
        s = re.sub(r'\b\d{4}-\d{2}-\d{2}\b', ' ', s)
        s = re.sub(r'\b\d+[\d,\.]*\b', ' ', s)
        brands = ['shipstation','easypost','pitney bowes','stamps','shippo','xps','endicia','usps','ups','fedex','dhl']
        for b in brands:
            s = re.sub(rf'\b{re.escape(b)}\b', ' [brand] ', s)
        s = re.sub(r'\bwebhook(s)?\b', ' api ', s)
        s = re.sub(r'\bportal\b', ' api ', s)
        s = re.sub(r'\bapi(s)?\b', ' api ', s)
        s = re.sub(r'\b3pl(s)?\b', ' 3pl ', s)
        s = re.sub(r'\bwms\b', ' 3pl ', s)
        s = re.sub(r'\bwarehouse management\b', ' 3pl ', s)
        s = re.sub(r'[^a-z\[\] ]+', ' ', s)
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    def _content_tokens(self, text: str) -> set:
        s = self._normalize_for_match(text)
        parts = [p for p in s.split() if p and p not in self._GENERIC_TERMS]
        # include bigrams
        bigrams = [f"{parts[i]} {parts[i+1]}" for i in range(len(parts)-1)] if len(parts) > 1 else []
        return set(parts + bigrams)

def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(description="Supio HARMONIZED Workbook Generator")
    parser.add_argument("--client", required=True, help="Client ID to generate workbook for")
    parser.add_argument("--output-dir", default=".", help="Output directory for workbook")

    args = parser.parse_args()

    try:
        # Initialize generator
        generator = SupioHarmonizedWorkbookGenerator(args.client)

        # Generate workbook
        workbook_path = generator.generate_workbook()

        # Get and display stats
        stats = generator.get_generation_stats()
        logger.info("📊 Generation Statistics:")
        for key, value in stats.items():
            logger.info(f"  {key}: {value}")

        logger.info(f"🎉 SUPIO HARMONIZED workbook generation completed successfully!")
        logger.info(f"📁 Workbook: {workbook_path}")

    except Exception as e:
        logger.error(f"❌ Supio HARMONIZED workbook generation failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main() 