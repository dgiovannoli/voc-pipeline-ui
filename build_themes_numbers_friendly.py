import os
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

SRC_WORKBOOK = "Demo Analyst Workbook.xlsx"
OUT_WORKBOOK = "Demo Analyst Workbook_THEMES_NUMBERS.xlsx"
THEMES_SHEET_NAME = "Themes"

WIN_SHEET_KEY = "win drivers"
SECTION_SHEET_MATCHERS = [
    ("win drivers", "Win Drivers Section"),
    ("loss factors", "Loss Factors Section"),
    ("competitive intelligence", "Competitive Intelligence"),
    ("implementation", "Implementation Insights"),
]

VALIDATION_OPTIONS = ["Validated", "Needs Revision", "Rejected", "Pending Review"]
SECTION_OPTIONS = [
    "Win Drivers",
    "Loss Factors",
    "Competitive Intelligence",
    "Implementation Insights",
    "Buyer Profile",
]


def find_sheet_by_key(wb, key: str) -> Worksheet:
    key = key.lower()
    for name in wb.sheetnames:
        if key in name.lower():
            return wb[name]
    raise RuntimeError(f"Sheet containing '{key}' not found")


def find_theme_blocks(ws: Worksheet) -> List[Tuple[int, int]]:
    blocks: List[Tuple[int, int]] = []
    start = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.startswith("theme_"):
            start = r
        if isinstance(a, str) and a.strip().upper().startswith("THEME DECISION:") and start is not None:
            blocks.append((start, r))
            start = None
    return blocks


def get_template_rows(ws_win: Worksheet) -> Tuple[int, int, int, int]:
    header_rows_end = 1
    header_row = None
    decision_row = None
    for r in range(1, ws_win.max_row + 1):
        a = ws_win.cell(row=r, column=1).value
        if isinstance(a, str) and a.startswith("theme_") and header_row is None:
            header_row = r
            break
        header_rows_end = r
    for r in range(header_row, ws_win.max_row + 1):
        a = ws_win.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().upper().startswith("THEME DECISION:"):
            decision_row = r
            break
    quote_row = header_row + 1
    return header_rows_end, header_row, quote_row, decision_row


def clone_style_from(src_cell, dst_cell):
    sfont = src_cell.font
    dst_cell.font = Font(name=sfont.name, size=sfont.size, bold=sfont.bold, italic=sfont.italic,
                         vertAlign=sfont.vertAlign, underline=sfont.underline, strike=sfont.strike,
                         color=sfont.color)
    sfill = src_cell.fill
    dst_cell.fill = PatternFill(fill_type=sfill.fill_type, fgColor=sfill.fgColor, bgColor=sfill.bgColor)
    salign = src_cell.alignment
    dst_cell.alignment = Alignment(horizontal=salign.horizontal, vertical=salign.vertical,
                                   wrap_text=salign.wrap_text, shrink_to_fit=salign.shrink_to_fit,
                                   indent=salign.indent, text_rotation=salign.text_rotation)
    sb = src_cell.border
    def copy_side(side: Side) -> Side:
        return Side(style=side.style, color=side.color)
    dst_cell.border = Border(left=copy_side(sb.left), right=copy_side(sb.right),
                             top=copy_side(sb.top), bottom=copy_side(sb.bottom))
    dst_cell.number_format = src_cell.number_format


def copy_row_style(src_ws: Worksheet, src_row: int, dst_ws: Worksheet, dst_row: int, max_col: int):
    for c in range(1, max_col + 1):
        clone_style_from(src_ws.cell(row=src_row, column=c), dst_ws.cell(row=dst_row, column=c))


def clear_data_after(ws: Worksheet, start_row: int):
    if ws.max_row > start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def unmerge_all(ws: Worksheet):
    ranges = list(ws.merged_cells.ranges)
    for r in ranges:
        try:
            ws.unmerge_cells(str(r))
        except Exception:
            pass


def apply_numbers_friendly_dropdowns(ws: Worksheet):
    # Clear any existing validations
    ws.data_validations.dataValidation = []
    val_csv = '"' + ",".join(VALIDATION_OPTIONS) + '"'
    sec_csv = '"' + ",".join(SECTION_OPTIONS) + '"'
    dv_val = DataValidation(type="list", formula1=val_csv, allow_blank=True, showDropDown=True)
    dv_sec = DataValidation(type="list", formula1=sec_csv, allow_blank=True, showDropDown=True)
    ws.add_data_validation(dv_val)
    ws.add_data_validation(dv_sec)

    applied = 0
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().upper().startswith("THEME DECISION:"):
            dv_val.add(ws.cell(row=r, column=14))  # N of decision row
            sec_row = r + 1
            dv_sec.add(ws.cell(row=sec_row, column=14))  # N of row below
            if not ws.cell(row=sec_row, column=2).value:
                ws.cell(row=sec_row, column=2, value="Report Section:").font = Font(bold=True)
            applied += 1
    return applied


def main():
    if not os.path.exists(SRC_WORKBOOK):
        raise FileNotFoundError(SRC_WORKBOOK)
    wb = load_workbook(SRC_WORKBOOK)

    ws_win = find_sheet_by_key(wb, WIN_SHEET_KEY)
    header_end, tmpl_header_row, tmpl_quote_row, tmpl_decision_row = get_template_rows(ws_win)
    max_col = ws_win.max_column

    if THEMES_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[THEMES_SHEET_NAME])
    ws_out = wb.copy_worksheet(ws_win)
    ws_out.title = THEMES_SHEET_NAME

    clear_data_after(ws_out, header_end + 1)

    # Append all blocks
    current = header_end + 1
    for key, _ in SECTION_SHEET_MATCHERS:
        try:
            src_ws = find_sheet_by_key(wb, key)
        except RuntimeError:
            continue
        for (start, end) in find_theme_blocks(src_ws):
            ws_out.insert_rows(current)
            copy_row_style(ws_win, tmpl_header_row, ws_out, current, max_col)
            for c in range(1, max_col + 1):
                ws_out.cell(row=current, column=c).value = src_ws.cell(row=start, column=c).value
            current += 1
            for r in range(start + 1, end):
                ws_out.insert_rows(current)
                copy_row_style(ws_win, tmpl_quote_row, ws_out, current, max_col)
                for c in range(1, max_col + 1):
                    ws_out.cell(row=current, column=c).value = src_ws.cell(row=r, column=c).value
                current += 1
            ws_out.insert_rows(current)
            copy_row_style(ws_win, tmpl_decision_row, ws_out, current, max_col)
            for c in range(1, max_col + 1):
                ws_out.cell(row=current, column=c).value = src_ws.cell(row=end, column=c).value
            current += 1
            # Spacer row
            ws_out.insert_rows(current)
            copy_row_style(ws_win, tmpl_quote_row, ws_out, current, max_col)
            for c in range(1, max_col + 1):
                ws_out.cell(row=current, column=c).value = None
            current += 1

    # Normalize merged cells and apply Numbers-friendly dropdowns
    unmerge_all(ws_out)
    applied = apply_numbers_friendly_dropdowns(ws_out)

    wb.save(OUT_WORKBOOK)
    print(f"✅ Rebuilt Numbers-friendly workbook with {applied} theme blocks → {OUT_WORKBOOK}")


if __name__ == "__main__":
    main() 