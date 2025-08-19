import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

IN_FILE = "Demo Analyst Workbook_THEMES_FROM_TEMPLATE.xlsx"
OUT_FILE = "Demo Analyst Workbook_THEMES_WITH_SECTION.xlsx"
SHEET_NAME = "Themes"

OPTIONS = [
    "Win Drivers",
    "Loss Factors",
    "Competitive Intelligence",
    "Implementation Insights",
    "Buyer Profile",
]


def main():
    if not os.path.exists(IN_FILE):
        raise FileNotFoundError(IN_FILE)
    wb = load_workbook(IN_FILE)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found")
    ws = wb[SHEET_NAME]

    # Determine the column to place the dropdown:
    # The Theme Decision value appears around column N (14) in the template, place dropdown at O (15)
    target_col = 15
    options_csv = ",".join(OPTIONS)
    dv = DataValidation(type="list", formula1=f'"{options_csv}"', allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)

    count = 0
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().upper().startswith("THEME DECISION:"):
            dv.add(ws.cell(row=r, column=target_col).coordinate)
            count += 1

    wb.save(OUT_FILE)
    print(f"✅ Added section dropdown to {count} decision rows → {OUT_FILE}")


if __name__ == "__main__":
    main() 