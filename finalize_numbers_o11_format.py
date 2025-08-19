import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

IN_FILE = "Demo Analyst Workbook_THEMES_NUMBERS_O11.xlsx"
OUT_FILE = "Demo Analyst Workbook_THEMES_NUMBERS_O11_FINAL.xlsx"
SHEET = "Themes"


def clone_style(src, dst):
    # Copy style attributes safely
    f = src.font
    dst.font = Font(name=f.name, size=f.size, bold=f.bold, italic=f.italic, vertAlign=f.vertAlign,
                    underline=f.underline, strike=f.strike, color=f.color)
    fill = src.fill
    dst.fill = PatternFill(fill_type=fill.fill_type, fgColor=fill.fgColor, bgColor=fill.bgColor)
    alg = src.alignment
    dst.alignment = Alignment(horizontal=alg.horizontal, vertical=alg.vertical, wrap_text=alg.wrap_text,
                              shrink_to_fit=alg.shrink_to_fit, indent=alg.indent, text_rotation=alg.text_rotation)
    b = src.border
    def cp(s: Side) -> Side:
        return Side(style=s.style, color=s.color)
    dst.border = Border(left=cp(b.left), right=cp(b.right), top=cp(b.top), bottom=cp(b.bottom))
    dst.number_format = src.number_format


def main():
    if not os.path.exists(IN_FILE):
        raise FileNotFoundError(IN_FILE)
    wb = load_workbook(IN_FILE)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET}' not found")
    ws = wb[SHEET]

    updated = 0
    cleared = 0

    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().upper().startswith("THEME DECISION:"):
            n11 = ws.cell(row=r, column=14)
            o11 = ws.cell(row=r, column=15)
            clone_style(n11, o11)
            # Clear N12 and its label in B of that row
            n12 = ws.cell(row=r + 1, column=14)
            n12.value = None
            # Best-effort style: set to default alignment with no wrap
            n12.alignment = Alignment(wrap_text=False)
            b12 = ws.cell(row=r + 1, column=2)
            if isinstance(b12.value, str) and b12.value.strip().lower().startswith("report section"):
                b12.value = None
            updated += 1
            cleared += 1

    wb.save(OUT_FILE)
    print(f"✅ Finalized formatting for {updated} O11 cells and cleared {cleared} N12 rows → {OUT_FILE}")


if __name__ == "__main__":
    main() 