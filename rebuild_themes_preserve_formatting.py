import os
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SRC_WORKBOOK = "Demo Analyst Workbook.xlsx"
OUT_WORKBOOK = "Demo Analyst Workbook_GROUPED.xlsx"
THEMES_SHEET_NAME = "Themes"

SECTION_SHEET_MATCHERS = [
    ("win drivers", "Win Drivers"),
    ("loss factors", "Loss Factors"),
    ("competitive intelligence", "Competitive Intelligence"),
    ("implementation", "Implementation Insights"),
]

FINAL_REPORT_SECTION_OPTIONS = [
    "Win Drivers",
    "Loss Factors",
    "Competitive Intelligence",
    "Implementation Insights",
    "Buyer Profile",
]


def find_section_sheets(wb) -> List[Tuple[str, str]]:
    results: List[Tuple[str, str]] = []  # (sheet_name, canonical_section)
    for sheet_name in wb.sheetnames:
        lower = sheet_name.strip().lower()
        for needle, canonical in SECTION_SHEET_MATCHERS:
            if needle in lower:
                results.append((sheet_name, canonical))
                break
    return results


def get_template_sheet_name(section_sheets: List[Tuple[str, str]]) -> str:
    for sname, _ in section_sheets:
        if "win drivers" in sname.lower():
            return sname
    return section_sheets[0][0]


def clear_values_only(ws: Worksheet, start_row: int, end_row: int, max_col: int):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c, value=None)


def ensure_column_header(ws: Worksheet, header_name: str, max_col: int) -> int:
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, max_col + 1)}
    if header_name in header_map and header_map[header_name]:
        return header_map[header_name]
    ws.cell(row=1, column=max_col + 1, value=header_name)
    return max_col + 1


def row_is_empty(ws: Worksheet, row_idx: int, last_col: int) -> bool:
    for c in range(1, last_col + 1):
        if ws.cell(row=row_idx, column=c).value not in (None, ""):
            return False
    return True


def append_rows_with_section(src_ws: Worksheet, dst_ws: Worksheet, section_label: str, final_section_col_idx: int, start_dst_row: int) -> int:
    rows_copied = 0
    dst_row = start_dst_row
    src_max_cols = src_ws.max_column
    for r in range(2, src_ws.max_row + 1):
        # Preserve structure: copy even blank separator rows, but skip truly empty rows across all columns
        if row_is_empty(src_ws, r, src_max_cols):
            continue
        for c in range(1, src_max_cols + 1):
            dst_ws.cell(row=dst_row, column=c, value=src_ws.cell(row=r, column=c).value)
        dst_ws.cell(row=dst_row, column=final_section_col_idx, value=section_label)
        rows_copied += 1
        dst_row += 1
    # Add a blank separator row between sections
    dst_row += 1
    return rows_copied


def add_final_section_validation(ws: Worksheet, final_section_col_idx: int, last_row: int):
    if not ws.cell(row=1, column=final_section_col_idx).value:
        ws.cell(row=1, column=final_section_col_idx, value="Final Report Section")
    options_csv = ",".join(FINAL_REPORT_SECTION_OPTIONS)
    dv = DataValidation(type="list", formula1=f'"{options_csv}"', allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    if last_row < 2:
        return
    col_letter = get_column_letter(final_section_col_idx)
    dv.add(f"{col_letter}2:{col_letter}{last_row}")


def expand_tables_and_filters(ws: Worksheet, last_row: int, last_col: int):
    last_col_letter = get_column_letter(last_col)
    ref = f"A1:{last_col_letter}{last_row}"
    # Update tables (if any)
    tables = list(getattr(ws, "_tables", []))
    for tbl in tables:
        try:
            tbl.ref = ref
        except Exception:
            pass
    # Update autofilter
    try:
        ws.auto_filter.ref = ref
    except Exception:
        pass


def main():
    if not os.path.exists(SRC_WORKBOOK):
        raise FileNotFoundError(SRC_WORKBOOK)
    wb = load_workbook(SRC_WORKBOOK, data_only=False)

    section_sheets = find_section_sheets(wb)
    if not section_sheets:
        raise RuntimeError("No section sheets found.")

    template_name = get_template_sheet_name(section_sheets)
    template_ws = wb[template_name]

    if THEMES_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[THEMES_SHEET_NAME])

    themes_ws = wb.copy_worksheet(template_ws)
    themes_ws.title = THEMES_SHEET_NAME

    template_max_cols = template_ws.max_column
    clear_values_only(themes_ws, start_row=2, end_row=max(2, themes_ws.max_row), max_col=template_max_cols)

    final_section_col_idx = ensure_column_header(themes_ws, "Final Report Section", template_max_cols)

    current_dst_row = 2
    total_rows = 0
    max_used_col = max(template_max_cols, final_section_col_idx)
    for sheet_name, canonical in section_sheets:
        src_ws = wb[sheet_name]
        copied = append_rows_with_section(src_ws, themes_ws, canonical, final_section_col_idx, current_dst_row)
        total_rows += copied
        current_dst_row += copied + 1  # leave a blank separator row
        if src_ws.max_column > max_used_col:
            max_used_col = src_ws.max_column

    last_row = current_dst_row - 2 if current_dst_row > 2 else 1

    add_final_section_validation(themes_ws, final_section_col_idx, last_row)
    expand_tables_and_filters(themes_ws, last_row=last_row, last_col=max_used_col)

    wb.save(OUT_WORKBOOK)
    print(f"✅ Wrote grouped workbook: {OUT_WORKBOOK}")
    print(f" - Themes rows: {total_rows}")
    print(f" - Final Report Section options: {', '.join(FINAL_REPORT_SECTION_OPTIONS)}")


if __name__ == "__main__":
    main() 