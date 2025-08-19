import os
import shutil
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

WORKBOOK_PATH = "Demo Analyst Workbook.xlsx"
BACKUP_PATH = WORKBOOK_PATH.replace(".xlsx", "_backup.xlsx")
THEMES_SHEET_NAME = "Themes"

# Map likely sheet name substrings to canonical section labels
SECTION_SHEET_MATCHERS = [
    ("win drivers", "Win Drivers"),
    ("loss factors", "Loss Factors"),
    ("competitive intelligence", "Competitive Intelligence"),
    ("implementation", "Implementation Insights"),
]

FINAL_REPORT_SECTION_OPTIONS = [
    "Win Drivers",
    "Loss Factors",
    "Competitive Intelligence",
    "Implementation Insights",
    "Buyer Profile",
]


def find_section_sheets(wb) -> List[Tuple[str, str]]:
    results: List[Tuple[str, str]] = []  # (sheet_name, canonical_section)
    for sheet_name in wb.sheetnames:
        lower = sheet_name.strip().lower()
        for needle, canonical in SECTION_SHEET_MATCHERS:
            if needle in lower:
                results.append((sheet_name, canonical))
                break
    return results


def get_or_append_column(ws: Worksheet, header_name: str) -> int:
    # Find header row as row 1
    max_col = ws.max_column
    existing = {ws.cell(row=1, column=c).value: c for c in range(1, max_col + 1)}
    if header_name in existing and existing[header_name]:
        return existing[header_name]
    ws.cell(row=1, column=max_col + 1, value=header_name)
    return max_col + 1


def clear_data_rows(ws: Worksheet):
    if ws.max_row <= 1:
        return
    ws.delete_rows(idx=2, amount=ws.max_row - 1)


def append_rows_with_section(src_ws: Worksheet, dst_ws: Worksheet, section_label: str, final_section_col_idx: int, max_copy_cols: int) -> int:
    rows_copied = 0
    for r in range(2, src_ws.max_row + 1):
        if all((src_ws.cell(row=r, column=c).value in (None, "")) for c in range(1, 4)):
            continue
        dst_row = dst_ws.max_row + 1
        for c in range(1, max_copy_cols + 1):
            dst_ws.cell(row=dst_row, column=c, value=src_ws.cell(row=r, column=c).value)
        dst_ws.cell(row=dst_row, column=final_section_col_idx, value=section_label)
        rows_copied += 1
    return rows_copied


def add_final_section_validation(ws: Worksheet, final_section_col_idx: int):
    # Ensure header
    if not ws.cell(row=1, column=final_section_col_idx).value:
        ws.cell(row=1, column=final_section_col_idx, value="Final Report Section")
    # Data validation list
    options_csv = ",".join(FINAL_REPORT_SECTION_OPTIONS)
    dv = DataValidation(type="list", formula1=f'"{options_csv}"', allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    max_row = max(ws.max_row, 1000)
    col_letter = ws.cell(row=1, column=final_section_col_idx).column_letter
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


def restore_backup_if_exists():
    if os.path.exists(BACKUP_PATH):
        shutil.copyfile(BACKUP_PATH, WORKBOOK_PATH)


def main():
    if not os.path.exists(WORKBOOK_PATH):
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    # Create backup if not exists
    if not os.path.exists(BACKUP_PATH):
        shutil.copyfile(WORKBOOK_PATH, BACKUP_PATH)
    else:
        # Restore from backup to get a clean base
        restore_backup_if_exists()

    wb = load_workbook(WORKBOOK_PATH)
    section_sheets = find_section_sheets(wb)
    if not section_sheets:
        raise RuntimeError("No section sheets found. Expected sheets like 'Win Drivers Section', 'Loss Factors Section', etc.")

    # Choose template sheet (Win Drivers preferred)
    template_name = None
    for sname, _ in section_sheets:
        if "win drivers" in sname.lower():
            template_name = sname
            break
    if template_name is None:
        template_name = section_sheets[0][0]

    # Remove existing Themes sheet if present
    if THEMES_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[THEMES_SHEET_NAME])

    # Duplicate template to preserve formatting
    template_ws = wb[template_name]
    themes_ws = wb.copy_worksheet(template_ws)
    themes_ws.title = THEMES_SHEET_NAME

    # Clear data rows in the duplicated sheet
    clear_data_rows(themes_ws)

    # Determine source columns to copy (use template header width)
    max_copy_cols = template_ws.max_column

    # Ensure Final Report Section column exists and get its index
    final_section_col_idx = get_or_append_column(themes_ws, "Final Report Section")

    # Append rows from each section
    total_rows = 0
    for sheet_name, canonical in section_sheets:
        src_ws = wb[sheet_name]
        rows = append_rows_with_section(src_ws, themes_ws, canonical, final_section_col_idx, max_copy_cols)
        total_rows += rows

    # Add data validation for Final Report Section including Buyer Profile
    add_final_section_validation(themes_ws, final_section_col_idx)

    wb.save(WORKBOOK_PATH)
    print(f"✅ Recreated '{WORKBOOK_PATH}' from backup and rebuilt '{THEMES_SHEET_NAME}' with {total_rows} rows.")
    print(f" - Final Report Section options: {', '.join(FINAL_REPORT_SECTION_OPTIONS)}")


if __name__ == "__main__":
    main() 