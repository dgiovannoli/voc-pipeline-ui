#!/usr/bin/env python3
import argparse
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

from win_loss_report_generator import WinLossReportGenerator
import yaml


def autosize(ws, max_col: int):
	widths = {}
	for row in ws.iter_rows(min_row=1, max_col=max_col):
		for cell in row:
			if cell.value is None:
				continue
			val = str(cell.value)
			widths[cell.column] = max(widths.get(cell.column, 0), len(val))
	for col, w in widths.items():
		ws.column_dimensions[get_column_letter(col)].width = min(max(w + 2, 10), 80)


def load_harmonization_patterns(config_path: Path):
	patterns = {}
	with open(config_path, 'r') as f:
		cfg = yaml.safe_load(f) or {}
	for subject, spec in (cfg.get('harmonization_patterns') or {}).items():
		kw = set(x.lower() for x in (spec.get('keywords') or []))
		patterns[subject] = kw
	return patterns


def suggest_subject_for_theme(theme: dict, patterns: dict) -> tuple:
	text = f"{theme.get('theme_statement','')}\n{theme.get('research_primary_question','')}".lower()
	tokens = set(t.strip('.,:;!?()[]"\'') for t in text.split() if t)
	best = (None, 0)
	for subject, kws in patterns.items():
		phrase_hits = sum(1 for k in kws if k in text)
		token_hits = sum(1 for k in kws if k in tokens)
		score = phrase_hits + 0.5 * token_hits
		if score > best[1]:
			best = (subject, score)
	conf = min(1.0, best[1] / 5.0) if best[1] > 0 else 0.0
	return best[0], conf


def _make_headline(raw: str) -> str:
	s = (raw or '').strip()
	for sep in ['. ', ' — ', ' - ', '; ']:
		if sep in s:
			return s.split(sep, 1)[0].strip()
	return s


def build_merged_themes(gen: WinLossReportGenerator):
	report = gen.generate_analyst_report()
	assert report.get('success'), f"Theme generation failed: {report.get('error')}"
	generated = report.get('themes', [])
	gen_by_id = {str(t.get('theme_id')): t for t in generated if t.get('theme_id')}

	# DB-backed research themes
	db = gen.db
	try:
		res = db.supabase.table('research_themes').select('*').eq('client_id', gen.client_id).eq('origin', 'research').execute()
		research_db = res.data or []
	except Exception:
		research_db = []
	# Evidence for counts
	theme_ids = [str(t.get('theme_id')) for t in research_db if t.get('theme_id')]
	evidence = []
	if theme_ids:
		try:
			res_ev = db.supabase.table('research_theme_evidence').select('*').in_('theme_id', theme_ids).execute()
			evidence = res_ev.data or []
		except Exception:
			try:
				res_ev2 = db.supabase.table('research_theme_evidence').select('*').in_('research_theme_id', theme_ids).execute()
				evidence = res_ev2.data or []
			except Exception:
				evidence = []
	# Response -> company map
	try:
		df = db.get_stage1_data_responses(client_id=gen.client_id)
		resp_to_company = {str(r['response_id']): (r.get('company') or r.get('company_name') or '') for _, r in df.iterrows()}
	except Exception:
		resp_to_company = {}
	# Counts
	ev_by_theme = defaultdict(list)
	for ev in evidence:
		tid = str(ev.get('theme_id') or ev.get('research_theme_id') or '')
		rid = str(ev.get('response_id') or ev.get('response_uuid') or ev.get('stage1_response_id') or '')
		if tid and rid:
			ev_by_theme[tid].append(rid)

	merged = list(generated)
	for t in research_db:
		tid = str(t.get('theme_id') or '')
		if not tid or tid in gen_by_id:
			continue
		metrics = {
			'quotes_count': len(ev_by_theme.get(tid, [])),
			'companies_count': len({resp_to_company.get(rid, '') for rid in ev_by_theme.get(tid, []) if resp_to_company.get(rid, '')}),
			'quality_score': float(t.get('quality_score', 0.0) or 0.0),
		}
		merged.append({
			'theme_id': tid,
			'theme_statement': t.get('theme_statement', ''),
			'theme_type': (t.get('theme_type') or 'research').lower(),
			'harmonized_subject': t.get('harmonized_subject') or '',
			'theme_origin': 'research',
			'research_primary_question': t.get('question_text') or t.get('primary_research_question') or '',
			'validation_metrics': metrics,
		})
	return merged


def add_all_themes_tab(xlsx_path: Path, client_id: str):
	gen = WinLossReportGenerator(client_id)
	themes = build_merged_themes(gen)
	patterns = load_harmonization_patterns(Path('config/subject_harmonization.yaml'))

	wb = load_workbook(filename=str(xlsx_path))

	# Compute research counts by scanning Research Themes sheet (display Theme ID in col C, Company in col G)
	research_quotes_by_display = defaultdict(int)
	research_companies_by_display = defaultdict(set)
	research_counts_by_headline = defaultdict(lambda: {'q':0,'companies':set()})
	if 'Research Themes' in wb.sheetnames:
		ws_r = wb['Research Themes']
		for r in range(6, ws_r.max_row + 1):
			id_val = ws_r.cell(r, 3).value  # Theme ID (display)
			company_val = ws_r.cell(r, 7).value  # Company
			label_val = ws_r.cell(r, 2).value  # Theme (Research) label
			if id_val:
				id_str = str(id_val)
				research_quotes_by_display[id_str] += 1
				if company_val:
					research_companies_by_display[id_str].add(str(company_val))
			# Headline-based counting
			if isinstance(label_val, str) and label_val.strip():
				base = label_val.split(' [', 1)[0].strip()
				entry = research_counts_by_headline[base]
				entry['q'] += 1
				if company_val:
					entry['companies'].add(str(company_val) )

	# Prefer DB-backed research metrics by raw theme_id
	research_by_tid = {}
	for t in themes:
		if (t.get('theme_origin','discovered') or 'discovered').lower() == 'research':
			rtid = str(t.get('theme_id') or '')
			vm = t.get('validation_metrics') or {}
			if rtid and vm:
				research_by_tid[rtid] = t

	# Build display IDs for research/discovered themes
	research_display_id_by_tid = {}
	try:
		res = gen.db.supabase.table('research_themes').select('*').eq('client_id', gen.client_id).eq('origin', 'research').execute()
		research_rows = res.data or []
	except Exception:
		research_rows = [t for t in themes if (t.get('theme_origin', 'discovered') or 'discovered').lower() == 'research']
	for idx, t in enumerate(research_rows, start=1):
		tid = str(t.get('theme_id') or '')
		if tid and tid not in research_display_id_by_tid:
			research_display_id_by_tid[tid] = f"research_theme_{idx:03d}"

	discovered_display_id_by_tid = {}
	discovered_rows = [t for t in themes if (t.get('theme_origin', 'discovered') or 'discovered').lower() != 'research']
	discovered_rows_sorted = sorted(
		discovered_rows,
		key=lambda x: ((x.get('harmonized_subject') or ''), (x.get('theme_statement') or '')),
	)
	for idx, t in enumerate(discovered_rows_sorted, start=1):
		tid = str(t.get('theme_id') or '')
		if tid and tid not in discovered_display_id_by_tid:
			discovered_display_id_by_tid[tid] = f"discovered_theme_{idx:03d}"

	if 'All Themes' in wb.sheetnames:
		wb.remove(wb['All Themes'])
	ws = wb.create_sheet('All Themes')

	# Styles
	title = Font(bold=True, size=16, color="FFFFFF")
	title_fill = PatternFill("solid", fgColor="2F4F4F")
	sub_fill = PatternFill("solid", fgColor="B0C4DE")
	header_font = Font(bold=True, color="FFFFFF")
	header_fill = PatternFill("solid", fgColor="4682B4")
	center = Alignment(horizontal="center", vertical="center", wrap_text=True)
	wrap = Alignment(vertical="top", wrap_text=True)

	# Title
	ws.merge_cells('A1:H1')
	ws['A1'] = 'All Themes - Review'
	ws['A1'].font = title
	for c in ws['A1:H1'][0]:
		c.fill = title_fill
	ws.merge_cells('A2:H2')
	ws['A2'] = 'Grouped by Harmonized Subject; Research themes auto-mapped when missing.'
	for c in ws['A2:H2'][0]:
		c.fill = sub_fill

	headers = ["Group", "Theme ID", "Theme Statement", "Origin", "Type", "Quotes", "Companies", "Quality Score"]
	for col, h in enumerate(headers, start=1):
		cell = ws.cell(row=5, column=col, value=h)
		cell.font = header_font
		cell.alignment = center
		cell.fill = header_fill

	row = 6
	groups = defaultdict(list)
	for t in themes:
		origin = (t.get('theme_origin', 'discovered') or 'discovered').lower()
		subj = (t.get('harmonized_subject') or '').strip()
		if origin == 'research' and not subj:
			sugg, _ = suggest_subject_for_theme(t, patterns)
			subj = sugg or (t.get('research_primary_question') or '').strip() or 'Research (Ungrouped)'
		if not subj:
			subj = 'General'
		groups[subj].append(t)

	for subj in sorted(groups.keys(), key=lambda s: (s is None, str(s))):
		ws.cell(row=row, column=1, value=subj).fill = sub_fill
		row += 1
		def _sort_key(t):
			origin = (t.get('theme_origin','discovered') or 'discovered').lower()
			origin_rank = 0 if origin == 'research' else 1
			qscore = float(t.get('validation_metrics',{}).get('quality_score', 0.0) or 0.0)
			return (origin_rank, -qscore, (t.get('theme_statement') or ''))
		for t in sorted(groups[subj], key=_sort_key):
			vm = t.get('validation_metrics', {})
			origin_t = (t.get('theme_origin', 'discovered') or 'discovered').title()
			raw_tid = str(t.get('theme_id') or '')
			origin_lower = (t.get('theme_origin', 'discovered') or 'discovered').lower()
			if origin_lower == 'research':
				id_display = research_display_id_by_tid.get(raw_tid, '')
				# Pull counts from DB-backed metrics if available, else fall back to sheet-derived counts
				src = research_by_tid.get(raw_tid)
				if src and (src.get('validation_metrics') or {}):
					vm2 = src['validation_metrics']
					quotes_count = int(vm2.get('quotes_count', 0) or 0)
					companies_count = int(vm2.get('companies_count', 0) or 0)
				else:
					# Try headline-based mapping when ID is missing in the sheet
					base = _make_headline(t.get('theme_statement',''))
					counts = research_counts_by_headline.get(base, {'q':0,'companies':set()})
					quotes_count = counts['q'] or research_quotes_by_display.get(id_display, 0)
					companies_count = len(counts['companies']) or len(research_companies_by_display.get(id_display, set()))
			else:
				id_display = discovered_display_id_by_tid.get(raw_tid, '')
				quotes_count = int(vm.get('quotes_count', 0) or 0)
				companies_count = int(vm.get('companies_count', 0) or 0)
			vals = [
				subj,
				id_display,
				t.get('theme_statement',''),
				origin_t,
				(t.get('theme_type','') or '').title(),
				quotes_count,
				companies_count,
				vm.get('quality_score',0.0)
			]
			for col, v in enumerate(vals, start=1):
				cell = ws.cell(row=row, column=col, value=v)
				cell.alignment = wrap if col in (1,3) else center
			row += 1

	autosize(ws, max_col=8)
	wb.save(str(xlsx_path))


def main():
	parser = argparse.ArgumentParser(description='Append an improved All Themes tab with subject suggestions')
	parser.add_argument('--client', required=True)
	parser.add_argument('--file', required=True)
	args = parser.parse_args()
	add_all_themes_tab(Path(args.file), args.client)
	print(f"✅ Added improved All Themes tab to {args.file}")

if __name__ == '__main__':
	main() 