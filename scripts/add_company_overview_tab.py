#!/usr/bin/env python3
import argparse
from collections import defaultdict, Counter
from pathlib import Path
from typing import Dict, Any, List

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

from win_loss_report_generator import WinLossReportGenerator


def autosize(ws, max_col: int):
	widths = {}
	for row in ws.iter_rows(min_row=1, max_col=max_col):
		for cell in row:
			if cell.value is None:
				continue
			val = str(cell.value)
			widths[cell.column] = max(widths.get(cell.column, 0), len(val))
	for col, w in widths.items():
		ws.column_dimensions[get_column_letter(col)].width = min(max(w + 2, 10), 80)


def find_raw_data_anchor(wb, company: str) -> int:
	"""Return 1-based row index of first occurrence of company in Raw Data sheet column B."""
	for sheet_name in ['📋 Raw Data', 'Raw Data', 'RAW DATA']:
		if sheet_name in wb.sheetnames:
			raw = wb[sheet_name]
			# headers at row 3 (1-based), data starts row 4
			for r in range(4, raw.max_row + 1):
				if (raw.cell(row=r, column=2).value or '').strip() == company:
					return r
	return 0


def add_company_overview_tab(xlsx_path: Path, client_id: str):
	gen = WinLossReportGenerator(client_id)
	try:
		df = gen.db.get_stage1_data_responses(client_id=client_id)
	except Exception:
		df = None
	records: List[Dict[str, Any]] = df.to_dict('records') if df is not None else []
	by_company: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
	for r in records:
		company = r.get('company') or r.get('company_name') or 'Unknown'
		by_company[company].append(r)

	wb = load_workbook(filename=str(xlsx_path))
	if 'Company Overview' in wb.sheetnames:
		wb.remove(wb['Company Overview'])
	ws = wb.create_sheet('Company Overview')

	# Styles
	title = Font(bold=True, size=16, color="FFFFFF")
	title_fill = PatternFill("solid", fgColor="2F4F4F")
	sub_fill = PatternFill("solid", fgColor="B0C4DE")
	header_font = Font(bold=True, color="FFFFFF")
	header_fill = PatternFill("solid", fgColor="4682B4")
	center = Alignment(horizontal="center", vertical="center", wrap_text=True)
	wrap = Alignment(vertical="top", wrap_text=True)

	# Title
	ws.merge_cells('A1:K1')
	ws['A1'] = 'Company Overview - Interview Context'
	ws['A1'].font = title
	for c in ws['A1:K1'][0]:
		c.fill = title_fill
	ws.merge_cells('A2:K2')
	ws['A2'] = 'Per-company snapshot: interviews, roles/titles, sentiment, deal mix, top subjects/questions, and quick link to Raw Data.'
	for c in ws['A2:K2'][0]:
		c.fill = sub_fill

	headers = [
		'Company', 'Interviews', 'Date Range', 'Roles/Titles', 'Industry', 'Deal Mix',
		'Sentiment Mix', 'Avg Impact', 'Top Subjects', 'Top Questions', 'Open Raw Data'
	]
	for col, h in enumerate(headers, start=1):
		cell = ws.cell(row=5, column=col, value=h)
		cell.font = header_font
		cell.alignment = center
		cell.fill = header_fill

	# Column widths
	ws.column_dimensions['A'].width = 28
	ws.column_dimensions['B'].width = 10
	ws.column_dimensions['C'].width = 24
	ws.column_dimensions['D'].width = 30
	ws.column_dimensions['E'].width = 20
	ws.column_dimensions['F'].width = 22
	ws.column_dimensions['G'].width = 22
	ws.column_dimensions['H'].width = 12
	ws.column_dimensions['I'].width = 28
	ws.column_dimensions['J'].width = 28
	ws.column_dimensions['K'].width = 20

	row = 6
	for company, group in sorted(by_company.items(), key=lambda kv: kv[0] or ''):
		interviews = len(group)
		dates = [r.get('created_at') for r in group if r.get('created_at')]
		try:
			date_min = min(dates) if dates else ''
			date_max = max(dates) if dates else ''
		except Exception:
			date_min = date_max = ''
		date_range = f"{date_min} — {date_max}" if (date_min or date_max) else ''

		roles = []
		for r in group:
			role = r.get('interviewee_title') or r.get('interviewee_role') or ''
			if not role:
				name = r.get('interviewee_name') or ''
				if ',' in name:
					parts = [p.strip() for p in name.split(',')]
					if len(parts) > 1:
						role = parts[1]
			if role:
				roles.append(role)
		roles_str = ', '.join(sorted(set(roles)))[:140]

		industry = ''
		for r in group:
			industry = r.get('industry') or r.get('client_industry') or ''
			if industry:
				break

		deal_counts = Counter((r.get('deal_status') or r.get('deal_outcome') or 'unknown').title() for r in group)
		deal_mix = ', '.join(f"{k}:{v}" for k, v in deal_counts.items())

		sent_counts = Counter((r.get('sentiment') or 'unknown').title() for r in group)
		sent_mix = ', '.join(f"{k}:{v}" for k, v in sent_counts.items())

		impacts = []
		for r in group:
			try:
				impacts.append(float(r.get('impact_score', 0) or 0))
			except Exception:
				pass
		avg_impact = sum(impacts) / len(impacts) if impacts else 0.0

		subjects = Counter((r.get('harmonized_subject') or '').strip() for r in group if (r.get('harmonized_subject') or '').strip())
		top_subjects = ', '.join(s for s, _ in subjects.most_common(3))

		questions = Counter((r.get('question') or '').strip() for r in group if (r.get('question') or '').strip())
		top_questions = ', '.join(q for q, _ in questions.most_common(3))

		ws.cell(row=row, column=1, value=company).alignment = wrap
		ws.cell(row=row, column=2, value=interviews).alignment = center
		ws.cell(row=row, column=3, value=date_range).alignment = center
		ws.cell(row=row, column=4, value=roles_str).alignment = wrap
		ws.cell(row=row, column=5, value=industry).alignment = center
		ws.cell(row=row, column=6, value=deal_mix).alignment = center
		ws.cell(row=row, column=7, value=sent_mix).alignment = center
		ws.cell(row=row, column=8, value=round(avg_impact, 1)).alignment = center
		ws.cell(row=row, column=9, value=top_subjects).alignment = wrap
		ws.cell(row=row, column=10, value=top_questions).alignment = wrap

		anchor_row = find_raw_data_anchor(wb, company)
		if anchor_row:
			# internal link to first row for the company in Raw Data
			for sheet_name in ['📋 Raw Data', 'Raw Data', 'RAW DATA']:
				if sheet_name in wb.sheetnames:
					ws.cell(row=row, column=11).hyperlink = f"#{sheet_name}!A{anchor_row}"
					ws.cell(row=row, column=11, value='Open Raw Data').alignment = center
					break
		row += 1

	autosize(ws, max_col=11)
	wb.save(str(xlsx_path))


def main():
	parser = argparse.ArgumentParser(description='Append a Company Overview tab to an existing workbook')
	parser.add_argument('--client', required=True, help='Client ID (e.g., Supio)')
	parser.add_argument('--file', required=True, help='Path to .xlsx file')
	args = parser.parse_args()
	add_company_overview_tab(Path(args.file), args.client)
	print(f"✅ Added Company Overview tab to {args.file}")

if __name__ == '__main__':
	main() 