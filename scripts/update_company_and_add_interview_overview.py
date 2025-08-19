#!/usr/bin/env python3
import argparse
import os
from collections import defaultdict, Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

from win_loss_report_generator import WinLossReportGenerator

try:
	import openai
except Exception:
	openai = None


def autosize(ws, max_col: int):
	widths = {}
	for row in ws.iter_rows(min_row=1, max_col=max_col):
		for cell in row:
			if cell.value is None:
				continue
			val = str(cell.value)
			widths[cell.column] = max(widths.get(cell.column, 0), len(val))
	for col, w in widths.items():
		ws.column_dimensions[get_column_letter(col)].width = min(max(w + 2, 10), 80)


def get_ai_overview(company: str, stats: Dict[str, Any], samples: List[str]) -> str:
	api_key = os.getenv('OPENAI_API_KEY')
	if not api_key:
		return ''
	# Build prompt once for both client styles
	person = (stats.get('person_hint') or '').strip()
	iv_summaries = stats.get('interview_summaries') or []
	context_block = ''
	if iv_summaries:
		context_block = "\n\nContext from interview-level summaries:\n- " + "\n- ".join(s[:240] for s in iv_summaries[:3])
	prompt = (
		f"Write a concise overview focused on perceived strengths, weaknesses, and competitors for the interview feedback at {company}.\n"
		f"Use only three lines exactly in this format:\n"
		f"Strengths: <1–2 specific strengths in plain language>\n"
		f"Weaknesses: <1–2 specific weaknesses or gaps in plain language>\n"
		f"Competitors: <competitors mentioned or likely compared against; use names if present, else descriptors>.\n"
		f"Do not include numbers or statistics. Avoid generic phrasing. Be specific and readable.{context_block}"
	)
	messages = [
		{"role": "system", "content": "You are an analyst summarizing interview context for a win/loss study."},
		{"role": "user", "content": prompt + ("\n\nSample quotes:\n- " + "\n- ".join(samples[:3]) if samples else '')}
	]
	try:
		# Preferred: OpenAI client (v1)
		try:
			from openai import OpenAI
			client = OpenAI(api_key=api_key)
			resp = client.chat.completions.create(
				model="gpt-4o-mini",
				messages=messages,
				temperature=0.4,
				max_tokens=120,
			)
			content = getattr(resp.choices[0].message, 'content', '')
			return (content or '').strip()
		except Exception:
			# Fallback: legacy openai module interface
			try:
				import openai as openai_legacy
				openai_legacy.api_key = api_key
				resp = openai_legacy.ChatCompletion.create(
					model="gpt-4o-mini",
					messages=messages,
					temperature=0.4,
					max_tokens=120,
				)
				return (resp.choices[0].message['content'] or '').strip()
			except Exception:
				return ''
	except Exception:
		return ''


def build_company_groups(records: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
	by_company: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
	for r in records:
		company = r.get('company') or r.get('company_name') or 'Unknown'
		by_company[company].append(r)
	return by_company


def simplify_company_overview(wb_path: Path, by_company: Dict[str, List[Dict[str, Any]]], company_to_summaries: Dict[str, List[str]] = None):
	wb = load_workbook(filename=str(wb_path))
	if 'Company Overview' in wb.sheetnames:
		wb.remove(wb['Company Overview'])
	ws = wb.create_sheet('Company Overview')

	# Styles
	title = Font(bold=True, size=16, color="FFFFFF")
	title_fill = PatternFill("solid", fgColor="2F4F4F")
	header_font = Font(bold=True, color="FFFFFF")
	header_fill = PatternFill("solid", fgColor="4682B4")
	center = Alignment(horizontal="center", vertical="center", wrap_text=True)
	wrap = Alignment(vertical="top", wrap_text=True)

	# Title
	ws.merge_cells('A1:F1')
	ws['A1'] = 'Company Overview - Analyst Context'
	ws['A1'].font = title
	for c in ws['A1:F1'][0]:
		c.fill = title_fill
	ws.merge_cells('A2:F2')
	ws['A2'] = 'Quotes, Industry, Deal Mix, Sentiment Mix, AI Overview'

	headers = ['Company', 'Quotes', 'Industry', 'Deal Mix', 'Sentiment Mix', 'AI Overview']
	for col, h in enumerate(headers, start=1):
		cell = ws.cell(row=5, column=col, value=h)
		cell.font = header_font
		cell.alignment = center
		cell.fill = header_fill

	# Column widths
	ws.column_dimensions['A'].width = 28
	ws.column_dimensions['B'].width = 10
	ws.column_dimensions['C'].width = 20
	ws.column_dimensions['D'].width = 22
	ws.column_dimensions['E'].width = 22
	ws.column_dimensions['F'].width = 80

	row = 6
	for company, group in sorted(by_company.items(), key=lambda kv: kv[0] or ''):
		quotes_count = len(group)
		# Industry (first non-empty)
		industry = ''
		for r in group:
			industry = r.get('industry') or r.get('client_industry') or ''
			if industry:
				break
		deal_counts = Counter((r.get('deal_status') or r.get('deal_outcome') or 'unknown').title() for r in group)
		deal_mix = ', '.join(f"{k}:{v}" for k, v in deal_counts.items())
		sent_counts = Counter((r.get('sentiment') or 'unknown').title() for r in group)
		sent_mix = ', '.join(f"{k}:{v}" for k, v in sent_counts.items())
		subjects = Counter((r.get('harmonized_subject') or '').strip() for r in group if (r.get('harmonized_subject') or '').strip())
		top_subjects = [s for s, _ in subjects.most_common(3)]
		sample_quotes = [
			(r.get('verbatim_response') or '')[:180].replace('\n',' ').strip()
			for r in group if (r.get('verbatim_response') or '').strip()
		]
		person_hint = ''
		for r in group:
			name = (r.get('interviewee_name') or '').strip()
			title = (r.get('interviewee_title') or r.get('interviewee_role') or '').strip()
			if name or title:
				person_hint = f"{name}, {title}".strip(', ')
				break
		stats = {
			'deal_mix': deal_mix,
			'sent_mix': sent_mix,
			'top_subjects': top_subjects,
			'person_hint': person_hint,
			'interview_summaries': (company_to_summaries or {}).get(company, []),
		}
		ai_summary = get_ai_overview(company, stats, sample_quotes)

		ws.cell(row=row, column=1, value=company).alignment = wrap
		ws.cell(row=row, column=2, value=quotes_count).alignment = center
		ws.cell(row=row, column=3, value=industry).alignment = center
		ws.cell(row=row, column=4, value=deal_mix).alignment = center
		ws.cell(row=row, column=5, value=sent_mix).alignment = center
		ws.cell(row=row, column=6, value=ai_summary).alignment = wrap
		row += 1

	autosize(ws, max_col=6)
	wb.save(str(wb_path))


def add_interview_overview_tab(wb_path: Path, by_company: Dict[str, List[Dict[str, Any]]]):
	wb = load_workbook(filename=str(wb_path))
	if 'Interview Overview' in wb.sheetnames:
		wb.remove(wb['Interview Overview'])
	ws = wb.create_sheet('Interview Overview')

	# Styles
	title = Font(bold=True, size=16, color="FFFFFF")
	title_fill = PatternFill("solid", fgColor="2F4F4F")
	header_font = Font(bold=True, color="FFFFFF")
	header_fill = PatternFill("solid", fgColor="4682B4")
	center = Alignment(horizontal="center", vertical="center", wrap_text=True)
	wrap = Alignment(vertical="top", wrap_text=True)

	# Title
	ws.merge_cells('A1:I1')
	ws['A1'] = 'Interview Overview - One Row Per Interview'
	ws['A1'].font = title
	for c in ws['A1:I1'][0]:
		c.fill = title_fill
	ws.merge_cells('A2:I2')
	ws['A2'] = 'Company, Interviewee, Title, Date, Quotes, Sentiment Mix, Deal Outcome, Coverage'

	headers = ['Interview Key', 'Company', 'Interviewee', 'Title', 'Date', 'Quotes', 'Sentiment Mix', 'Deal Outcome', 'Distinct Questions']
	for col, h in enumerate(headers, start=1):
		cell = ws.cell(row=5, column=col, value=h)
		cell.font = header_font
		cell.alignment = center
		cell.fill = header_fill

	# Column widths
	ws.column_dimensions['A'].width = 24
	ws.column_dimensions['B'].width = 26
	ws.column_dimensions['C'].width = 26
	ws.column_dimensions['D'].width = 26
	ws.column_dimensions['E'].width = 16
	ws.column_dimensions['F'].width = 10
	ws.column_dimensions['G'].width = 22
	ws.column_dimensions['H'].width = 16
	ws.column_dimensions['I'].width = 18

	row = 6
	for company, group in sorted(by_company.items(), key=lambda kv: kv[0] or ''):
		# Group rows by interview key: prefer interview_id; else composite key
		by_interview: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
		for r in group:
			ikey = r.get('interview_id')
			if not ikey:
				date_str = ''
				raw_date = r.get('created_at') or ''
				if raw_date:
					date_str = str(raw_date)[:10]
				ikey = f"{company} | {r.get('interviewee_name','Unknown')} | {date_str}"
			by_interview[ikey].append(r)

		for ikey, items in sorted(by_interview.items(), key=lambda kv: kv[0]):
			quotes = len(items)
			interviewee = items[0].get('interviewee_name') or ''
			title = items[0].get('interviewee_title') or items[0].get('interviewee_role') or ''
			date_val = items[0].get('created_at') or ''
			date_disp = str(date_val)[:10] if date_val else ''
			sent_counts = Counter((r.get('sentiment') or 'unknown').title() for r in items)
			sent_mix = ', '.join(f"{k}:{v}" for k, v in sent_counts.items())
			deal = (items[0].get('deal_status') or items[0].get('deal_outcome') or '').title()
			questions = Counter((r.get('question') or '').strip() for r in items if (r.get('question') or '').strip())
			distinct_q = len(questions)

			ws.cell(row=row, column=1, value=ikey).alignment = wrap
			ws.cell(row=row, column=2, value=company).alignment = wrap
			ws.cell(row=row, column=3, value=interviewee).alignment = wrap
			ws.cell(row=row, column=4, value=title).alignment = wrap
			ws.cell(row=row, column=5, value=date_disp).alignment = center
			ws.cell(row=row, column=6, value=quotes).alignment = center
			ws.cell(row=row, column=7, value=sent_mix).alignment = center
			ws.cell(row=row, column=8, value=deal).alignment = center
			ws.cell(row=row, column=9, value=distinct_q).alignment = center
			row += 1

	autosize(ws, max_col=9)
	wb.save(str(wb_path))


def load_company_summaries(gen: WinLossReportGenerator) -> Dict[str, List[str]]:
	by_company: Dict[str, List[str]] = {}
	try:
		resp = gen.db.supabase.table('interview_summaries').select('*').eq('client_id', gen.client_id).execute()
		rows = resp.data or []
		for r in rows:
			comp = r.get('company') or 'Unknown'
			text = (r.get('summary') or '').strip()
			if text:
				by_company.setdefault(comp, []).append(text)
		return by_company
	except Exception:
		# CSV fallback
		try:
			from csv import DictReader
			csv_path = Path(f"interview_summaries_{gen.client_id}.csv")
			if csv_path.exists():
				with csv_path.open() as f:
					for r in DictReader(f):
						comp = r.get('company') or 'Unknown'
						text = (r.get('summary') or '').strip()
						if text:
							by_company.setdefault(comp, []).append(text)
			return by_company
		except Exception:
			return {}


def main():
	parser = argparse.ArgumentParser(description='Update Company Overview and add Interview Overview tabs')
	parser.add_argument('--client', required=True)
	parser.add_argument('--file', required=True)
	args = parser.parse_args()

	gen = WinLossReportGenerator(args.client)
	try:
		df = gen.db.get_stage1_data_responses(client_id=args.client)
	except Exception:
		df = None
	records = df.to_dict('records') if df is not None else []
	by_company = build_company_groups(records)

	# Update company overview
	company_summaries = load_company_summaries(gen)
	simplify_company_overview(Path(args.file), by_company, company_summaries)
	# Add interview overview
	add_interview_overview_tab(Path(args.file), by_company)
	print(f"✅ Updated Company Overview and added Interview Overview to {args.file}")

if __name__ == '__main__':
	main() 