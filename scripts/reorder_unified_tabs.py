#!/usr/bin/env python3
import argparse
from pathlib import Path
from openpyxl import load_workbook

# Desired order for unified flow
TARGET_ORDER = [
	"All Themes",
	"Research Themes",
	"Discovered Themes",
	"Mapping QA",
	"📋 Raw Data",
]


def reorder_sheets(xlsx_path: Path):
	wb = load_workbook(filename=str(xlsx_path))
	existing = list(wb.sheetnames)
	# Build new ordered list: target in order if exists, then any others afterwards
	ordered = [name for name in TARGET_ORDER if name in existing]
	ordered += [name for name in existing if name not in ordered]
	# Reorder in workbook
	for idx, name in enumerate(ordered):
		wb._sheets[idx] = wb[name]
	# Save
	wb.save(str(xlsx_path))


def main():
	p = argparse.ArgumentParser(description="Reorder workbook sheets to unified order")
	p.add_argument("--file", required=True)
	args = p.parse_args()
	reorder_sheets(Path(args.file))
	print(f"✅ Reordered sheets in {args.file}")

if __name__ == "__main__":
	main() 