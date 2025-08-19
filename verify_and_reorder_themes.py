import os
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

IN_FILE = "Demo Analyst Workbook_THEMES_NUMBERS_O11_FINAL.xlsx"
OUT_FILE = "Demo Analyst Workbook_FINAL.xlsx"

THEMES_SHEET = "Themes"
SECTION_SHEETS = [
    "Win Drivers Section",
    "Loss Factors Section",
    "Competitive Intelligence",
    "Implementation Insights",
    "Pricing Analysis",
]


def find_theme_blocks(ws: Worksheet) -> List[Tuple[int, int]]:
    blocks: List[Tuple[int, int]] = []
    start = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.startswith("theme_"):
            start = r
        if isinstance(a, str) and a.strip().upper().startswith("THEME DECISION:") and start is not None:
            blocks.append((start, r))
            start = None
    return blocks


def count_blocks_in_sheetnames(wb, names: List[str]) -> int:
    total = 0
    for name in names:
        if name in wb.sheetnames:
            total += len(find_theme_blocks(wb[name]))
    return total


def main():
    if not os.path.exists(IN_FILE):
        raise FileNotFoundError(IN_FILE)
    wb = load_workbook(IN_FILE)

    if THEMES_SHEET not in wb.sheetnames:
        raise RuntimeError("Themes sheet not found")

    themes_blocks = len(find_theme_blocks(wb[THEMES_SHEET]))
    section_blocks = count_blocks_in_sheetnames(wb, SECTION_SHEETS)

    print(f"Themes blocks: {themes_blocks} | Section blocks: {section_blocks}")

    # Proceed only if coverage matches or Themes has at least as many blocks
    if themes_blocks < section_blocks:
        raise RuntimeError("Themes sheet appears to be missing themes. Aborting reordering/deletions.")

    # Move Themes to be second tab (index 1)
    ws_themes = wb[THEMES_SHEET]
    # Calculate current position
    current_index = wb.sheetnames.index(THEMES_SHEET)
    # Move left by (current_index - 1)
    offset = 1 - current_index
    if offset != 0:
        wb.move_sheet(ws_themes, offset=offset)

    # Remove specified section sheets if present
    removed = []
    for name in SECTION_SHEETS:
        if name in wb.sheetnames:
            wb.remove(wb[name])
            removed.append(name)

    wb.save(OUT_FILE)
    print(f"✅ Saved {OUT_FILE}. Removed sheets: {', '.join(removed) if removed else 'none'}. Themes moved to second tab.")


if __name__ == "__main__":
    main() 