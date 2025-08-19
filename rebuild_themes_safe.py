import os
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

SRC_WORKBOOK = "Demo Analyst Workbook.xlsx"
OUT_WORKBOOK = "Demo Analyst Workbook_FIXED.xlsx"
THEMES_SHEET_NAME = "Themes"

SECTION_SHEET_MATCHERS = [
    ("win drivers", "Win Drivers"),
    ("loss factors", "Loss Factors"),
    ("competitive intelligence", "Competitive Intelligence"),
    ("implementation", "Implementation Insights"),
]

FINAL_REPORT_SECTION_OPTIONS = [
    "Win Drivers",
    "Loss Factors",
    "Competitive Intelligence",
    "Implementation Insights",
    "Buyer Profile",
]


def find_section_sheets(wb) -> List[Tuple[str, str]]:
    results: List[Tuple[str, str]] = []  # (sheet_name, canonical_section)
    for sheet_name in wb.sheetnames:
        lower = sheet_name.strip().lower()
        for needle, canonical in SECTION_SHEET_MATCHERS:
            if needle in lower:
                results.append((sheet_name, canonical))
                break
    return results


def copy_header(src_ws: Worksheet, dst_ws: Worksheet) -> int:
    max_col = src_ws.max_column
    for col in range(1, max_col + 1):
        dst_ws.cell(row=1, column=col, value=src_ws.cell(row=1, column=col).value)
    return max_col


def append_rows_with_section(src_ws: Worksheet, dst_ws: Worksheet, section_label: str, final_section_col_idx: int, max_copy_cols: int) -> int:
    rows_copied = 0
    for r in range(2, src_ws.max_row + 1):
        # Skip empty rows (first 3 cols)
        if all((src_ws.cell(row=r, column=c).value in (None, "")) for c in range(1, 4)):
            continue
        dst_row = dst_ws.max_row + 1
        for c in range(1, max_copy_cols + 1):
            dst_ws.cell(row=dst_row, column=c, value=src_ws.cell(row=r, column=c).value)
        dst_ws.cell(row=dst_row, column=final_section_col_idx, value=section_label)
        rows_copied += 1
    return rows_copied


def add_final_section_validation(ws: Worksheet, final_section_col_idx: int, last_row: int):
    if not ws.cell(row=1, column=final_section_col_idx).value:
        ws.cell(row=1, column=final_section_col_idx, value="Final Report Section")
    options_csv = ",".join(FINAL_REPORT_SECTION_OPTIONS)
    dv = DataValidation(type="list", formula1=f'"{options_csv}"', allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    if last_row < 2:
        return
    col_letter = ws.cell(row=1, column=final_section_col_idx).column_letter
    dv.add(f"{col_letter}2:{col_letter}{last_row}")


def main():
    if not os.path.exists(SRC_WORKBOOK):
        raise FileNotFoundError(SRC_WORKBOOK)
    wb = load_workbook(SRC_WORKBOOK, data_only=False)

    section_sheets = find_section_sheets(wb)
    if not section_sheets:
        raise RuntimeError("No section sheets found.")

    # Determine header template from Win Drivers if available
    header_src_name = None
    for sname, _ in section_sheets:
        if "win drivers" in sname.lower():
            header_src_name = sname
            break
    if header_src_name is None:
        header_src_name = section_sheets[0][0]

    # Remove existing Themes in output if copying in place
    if THEMES_SHEET_NAME in wb.sheetnames:
        ws_existing = wb[THEMES_SHEET_NAME]
        wb.remove(ws_existing)

    themes_ws = wb.create_sheet(THEMES_SHEET_NAME, 0)

    # Copy header row values
    header_ws = wb[header_src_name]
    max_copy_cols = copy_header(header_ws, themes_ws)

    # Ensure Final Report Section column exists and get its index
    # If exists in header, reuse index; else append
    header_map = {themes_ws.cell(row=1, column=c).value: c for c in range(1, max_copy_cols + 1)}
    if "Final Report Section" in header_map and header_map["Final Report Section"]:
        final_section_col_idx = header_map["Final Report Section"]
    else:
        final_section_col_idx = max_copy_cols + 1
        themes_ws.cell(row=1, column=final_section_col_idx, value="Final Report Section")

    # Append data from each section
    total_rows = 0
    for sheet_name, canonical in section_sheets:
        src_ws = wb[sheet_name]
        total_rows += append_rows_with_section(src_ws, themes_ws, canonical, final_section_col_idx, max_copy_cols)

    # Add data validation only to the used rows
    add_final_section_validation(themes_ws, final_section_col_idx, last_row=themes_ws.max_row)

    wb.save(OUT_WORKBOOK)
    print(f"✅ Wrote safe rebuilt workbook: {OUT_WORKBOOK}")
    print(f" - Unified sheet: '{THEMES_SHEET_NAME}' with {total_rows} rows")
    print(f" - Final Report Section includes: {', '.join(FINAL_REPORT_SECTION_OPTIONS)}")


if __name__ == "__main__":
    main() 