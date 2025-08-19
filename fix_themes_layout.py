import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

IN_FILE = "Demo Analyst Workbook_GROUPED.xlsx"
OUT_FILE = "Demo Analyst Workbook_CLEAN.xlsx"
SHEET_NAME = "Themes"

SECTION_OPTIONS = [
    "Win Drivers",
    "Loss Factors",
    "Competitive Intelligence",
    "Implementation Insights",
    "Buyer Profile",
]


def remove_column_p(ws):
    max_col = ws.max_column
    if max_col >= 16:
        ws.delete_cols(16, 1)


def unmerge_all(ws):
    ranges = list(ws.merged_cells.ranges)
    for r in ranges:
        try:
            ws.unmerge_cells(str(r))
        except Exception:
            pass


def reset_style(cell):
    cell.fill = PatternFill()
    cell.font = Font()
    cell.alignment = Alignment()
    cell.border = Border()


def normalize_theme_blocks(ws):
    current_block = False
    for r in range(2, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        if isinstance(a_val, str) and a_val.strip().upper().startswith("THEME DECISION:"):
            current_block = False
            continue
        if isinstance(a_val, str) and a_val.startswith("theme_"):
            current_block = True
            continue
        if current_block:
            for c in range(1, 6):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                reset_style(cell)


def add_section_dropdown_in_decision_rows(ws):
    section_col = 15
    options_csv = ",".join(SECTION_OPTIONS)
    dv = DataValidation(type="list", formula1=f'"{options_csv}"', allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    for r in range(2, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        if isinstance(a_val, str) and a_val.strip().upper().startswith("THEME DECISION:"):
            dv.add(ws.cell(row=r, column=section_col).coordinate)


def main():
    if not os.path.exists(IN_FILE):
        raise FileNotFoundError(IN_FILE)
    wb = load_workbook(IN_FILE)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found")
    ws = wb[SHEET_NAME]

    unmerge_all(ws)
    remove_column_p(ws)
    normalize_theme_blocks(ws)
    add_section_dropdown_in_decision_rows(ws)

    wb.save(OUT_FILE)
    print(f"✅ Wrote cleaned workbook: {OUT_FILE}")


if __name__ == "__main__":
    main() 