import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment

IN_FILE = "Demo Analyst Workbook_THEMES_NUMBERS_FINAL.xlsx"
OUT_FILE = "Demo Analyst Workbook_THEMES_NUMBERS_O11.xlsx"
SHEET = "Themes"

SECTION_OPTIONS = [
    "Win Drivers",
    "Loss Factors",
    "Competitive Intelligence",
    "Implementation Insights",
    "Buyer Profile",
]


def main():
    if not os.path.exists(IN_FILE):
        raise FileNotFoundError(IN_FILE)
    wb = load_workbook(IN_FILE)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET}' not found")
    ws = wb[SHEET]

    # Build Numbers-friendly inline CSV list
    sec_csv = '"' + ",".join(SECTION_OPTIONS) + '"'
    dv_sec = DataValidation(type="list", formula1=sec_csv, allow_blank=True, showDropDown=True)
    ws.add_data_validation(dv_sec)

    moved = 0
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().upper().startswith("THEME DECISION:"):
            # O on decision row (column 15)
            o_cell = ws.cell(row=r, column=15)
            # Default if empty so Numbers shows a pop-up
            if not o_cell.value:
                o_cell.value = SECTION_OPTIONS[0]
            dv_sec.add(o_cell)
            # Mirror into N on the next row (r+1, col 14)
            n12 = ws.cell(row=r + 1, column=14)
            n12.value = f"=O{r}"
            # Ensure wrapping is off just in case
            n12.alignment = Alignment(wrap_text=False)
            moved += 1

    wb.save(OUT_FILE)
    print(f"✅ Added Report Section dropdowns to O on {moved} decision rows and mirrored to N on the row below → {OUT_FILE}")


if __name__ == "__main__":
    main() 